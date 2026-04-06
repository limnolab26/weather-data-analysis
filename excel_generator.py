# excel_generator.py  v2.0
# 엑셀 보고서 생성 모듈
# 시트 구성:
#  1. 📋 보고서 요약
#  2. 📅 월별 통계
#  3. 📊 원본 데이터  (연도·월 열 추가)
#  4. 📈 기상 차트
#  5. 🔄 피벗 분석    (AVERAGEIFS·SUMIFS 수식 기반)
#  6. 🌧️ 강수량 분석
#  7. 🌡️ 기상 특성 검토  ← NEW
#  8. 🌍 기후변화 검토    ← NEW
#  9. 피벗작업            (실제 Excel PivotTable 삽입)
# 10. 누적강수량 분석

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import Marker
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import numpy as np
from datetime import datetime
import zipfile, io, re

from data_processor import ELEMENT_LABELS, SUM_ELEMENTS, MODE_ELEMENTS

# ── 색상 팔레트 ─────────────────────────────────────────────────────────────
C = {
    'dark_blue':    '1F4E79',
    'mid_blue':     '2E75B6',
    'light_blue':   'DEEAF1',
    'white':        'FFFFFF',
    'orange':       'C55A11',
    'light_orange': 'FCE4D6',
    'green':        '375623',
    'light_green':  'E2EFDA',
    'mid_green':    '70AD47',
    'gray':         '595959',
    'light_gray':   'F2F2F2',
    'border':       'BFBFBF',
    'yellow':       'FFE699',
    'light_yellow': 'FFF2CC',
    'teal':         '00B0F0',
    'light_teal':   'DDEBF7',
}
FONT = '맑은 고딕'
RAW_SHEET = "'📊 원본 데이터'"   # 수식에서 사용할 시트 참조


def _thin(color=C['border']):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _hc(ws, row, col, val, *, bg=C['dark_blue'], fg=C['white'],
        bold=True, sz=10, align='center', wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT, bold=bold, color=fg, size=sz)
    c.fill      = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border    = _thin()
    return c


def _dc(ws, row, col, val, *, bg=C['white'], bold=False, sz=9,
        align='center', nf=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT, bold=bold, size=sz, color='000000')
    c.fill      = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border    = _thin()
    if nf:
        c.number_format = nf
    return c


def _title(ws, row, cs, ce, text, h=24, bg=C['dark_blue'], sz=12):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=text)
    c.font      = Font(name=FONT, bold=True, size=sz, color=C['white'])
    c.fill      = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = h
    return c


def _note(ws, row, cs, ce, text):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=text)
    c.font      = Font(name=FONT, size=8, color=C['gray'])
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 13


def _set_chart_layout(chart, margin_ratio=0.08):
    """
    차트 그림 영역(plot area)에 여백을 설정합니다.
    margin_ratio: 차트 전체 대비 여백 비율 (10mm ≈ 0.08 기준)
    좌우·상하 각각 margin_ratio만큼 여백 적용.
    """
    ml = ManualLayout(
        layoutTarget='inner',
        xMode='edge', yMode='edge',
        wMode='edge', hMode='edge',
        x=margin_ratio,
        y=margin_ratio * 1.5,          # 상단은 제목 공간 감안해 조금 더
        w=1.0 - margin_ratio * 2,
        h=1.0 - margin_ratio * 3.5,    # 하단은 x축 레이블 공간 감안
    )
    chart.plot_area.layout = Layout(manualLayout=ml)


class ExcelReportGenerator:
    """엑셀 보고서 생성기 v2.0"""

    def __init__(self):
        # 원본 데이터 시트 컬럼 위치 추적 (수식 참조용)
        self._col = {}       # key → column letter (e.g. 'temp_avg' → 'E')
        self._n_rows = 0     # 데이터 행 수 (헤더 제외)
        self._years  = []    # 분석 연도 목록
        self._has_stn = False

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 메인 생성 함수
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def generate(self, df: pd.DataFrame, stats: dict, config: dict,
                 output_path: str, selected_cols: list) -> None:
        df2 = df.copy()
        df2['_y'] = df2['date'].dt.year
        self._years    = sorted(df2['_y'].unique().tolist())
        self._n_rows   = len(df)
        self._has_stn  = 'station_name' in df.columns

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._sheet_summary(wb, df, stats, config)
        self._sheet_monthly(wb, stats)
        self._sheet_raw(wb, df, selected_cols)        # ← col 맵핑 설정
        self._sheet_charts(wb, df, selected_cols)
        self._sheet_pivot(wb, df, selected_cols)      # ← 수식 기반
        self._sheet_precipitation(wb, df, selected_cols)
        self._sheet_weather_chars(wb, df, selected_cols)   # NEW
        self._sheet_climate_change(wb, df, selected_cols)  # NEW
        self._sheet_pivot_work(wb, df, selected_cols)      # PivotTable 자리
        self._sheet_cumulative_precip(wb, df, selected_cols)
        self._sheet_raw2(wb, df, selected_cols)          # ← NEW: 세로형 데이터
        self._sheet_rainfall_days(wb, df, selected_cols) # ← NEW: 강우일수
        self._sheet_monthly_tp(wb, df, selected_cols)    # ← NEW: 월기온&강수량
        # 피벗작업2 시트 틀 생성
        ws_pivot2 = wb.create_sheet('피벗작업2')
        _note(ws_pivot2, 1, 1, 8,
              '원본 데이터2(세로형) 기반 피벗테이블 — 클릭하면 필드 목록이 표시됩니다.')
        _note(ws_pivot2, 2, 1, 8,
              '※ 날짜·연도·월·관측소명·항목·Data 필드를 드래그해 자유롭게 분석하세요.')
        self._sheet_boxplot(wb, df, selected_cols)           # ← NEW: Box Plot

        wb.save(output_path)

        # Table _rels 경로 및 중복 autoFilter 수정
        self._fix_table_rels(output_path)
        # 차트 그림 영역 여백 삽입 (ZIP XML 직접 수정)
        self._inject_chart_layouts(output_path)

        # 실제 Excel PivotTable 삽입 (파일 저장 후)
        self._inject_pivot_table(output_path, df, selected_cols)
        self._inject_pivot_table2(output_path)   # ← NEW: 원본 데이터2 피벗

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 1. 보고서 요약
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_summary(self, wb, df, stats, config):
        ws = wb.create_sheet("📋 보고서 요약")
        ws.sheet_view.showGridLines = False
        title  = config.get('report_title', '기상 현황 보고서')
        org    = config.get('organization', '')
        today  = datetime.now().strftime('%Y년 %m월 %d일')
        s_date = df['date'].min().strftime('%Y-%m-%d')
        e_date = df['date'].max().strftime('%Y-%m-%d')

        ws.merge_cells('B2:L4')
        t = ws['B2']
        t.value     = title
        t.font      = Font(name=FONT, bold=True, size=22, color=C['dark_blue'])
        t.alignment = Alignment(horizontal='center', vertical='center')
        for r in [2, 3, 4]:
            ws.row_dimensions[r].height = 20

        ws.merge_cells('B5:L5')
        ws['B5'].fill = PatternFill(start_color=C['dark_blue'], fill_type='solid')
        ws.row_dimensions[5].height = 4

        ws.merge_cells('B6:L6')
        meta = f"분석 기간: {s_date} ~ {e_date}"
        if org:
            meta += f"   |   작성 기관: {org}"
        meta += f"   |   작성일: {today}"
        m = ws['B6']
        m.value     = meta
        m.font      = Font(name=FONT, size=10, color=C['gray'])
        m.alignment = Alignment(horizontal='center')
        ws.row_dimensions[6].height = 18

        row = 8
        for stn, stn_stats in stats.items():
            ws.merge_cells(f'B{row}:L{row}')
            c = ws[f'B{row}']
            c.value     = f"▶  관측소: {stn}"
            c.font      = Font(name=FONT, bold=True, size=12, color=C['white'])
            c.fill      = PatternFill(start_color=C['dark_blue'], fill_type='solid')
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 22
            row += 1

            period = stn_stats.get('period', {})
            sd, ed = period.get('start'), period.get('end')
            ws.merge_cells(f'B{row}:L{row}')
            sub = ws[f'B{row}']
            sub.value = (f"  관측 기간: "
                         f"{sd.strftime('%Y-%m-%d') if sd else '-'} ~ "
                         f"{ed.strftime('%Y-%m-%d') if ed else '-'}  "
                         f"(총 {period.get('days', 0):,}일)")
            sub.font      = Font(name=FONT, size=10, color=C['gray'])
            sub.alignment = Alignment(horizontal='left', indent=1)
            ws.row_dimensions[row].height = 16
            row += 1

            _hc(ws, row, 2, '기상요소', bg=C['mid_blue'], sz=10)
            _hc(ws, row, 4, '통계 항목', bg=C['mid_blue'], sz=10)
            _hc(ws, row, 6, '값',        bg=C['mid_blue'], sz=10)
            ws.merge_cells(f'B{row}:C{row}')
            ws.merge_cells(f'D{row}:E{row}')
            ws.merge_cells(f'F{row}:L{row}')
            ws.row_dimensions[row].height = 18
            row += 1

            overall = stn_stats.get('overall', {})
            for i, (elem, values) in enumerate(overall.items()):
                bg = C['light_blue'] if i % 2 == 0 else C['white']
                first = True
                for stat_name, stat_val in values.items():
                    ws.merge_cells(f'B{row}:C{row}')
                    ws.merge_cells(f'D{row}:E{row}')
                    ws.merge_cells(f'F{row}:L{row}')
                    if first:
                        _dc(ws, row, 2, elem, bg=bg, bold=True, sz=10)
                        first = False
                    else:
                        _dc(ws, row, 2, '', bg=bg)
                    _dc(ws, row, 4, stat_name, bg=bg, sz=10)
                    _dc(ws, row, 6, stat_val,  bg=bg, sz=10)
                    ws.row_dimensions[row].height = 16
                    row += 1
            row += 2

        for col, w in {'A':2,'B':3,'C':18,'D':3,'E':12,'F':3,
                       'G':10,'H':10,'I':10,'J':10,'K':10,'L':10}.items():
            ws.column_dimensions[col].width = w

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 2. 월별 통계
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_monthly(self, wb, stats):
        ws = wb.create_sheet("📅 월별 통계")
        ws.sheet_view.showGridLines = False
        row = 1
        for stn, stn_stats in stats.items():
            monthly_all = stn_stats.get('monthly', {})
            if not monthly_all:
                continue
            ws.merge_cells(f'A{row}:AZ{row}')
            c = ws[f'A{row}']
            c.value     = f"관측소: {stn}  —  월별 기상 통계"
            c.font      = Font(name=FONT, bold=True, size=12, color=C['white'])
            c.fill      = PatternFill(start_color=C['dark_blue'], fill_type='solid')
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 22
            row += 2

            for elem_label, monthly_data in monthly_all.items():
                if not monthly_data:
                    continue
                ws[f'A{row}'].value     = f"【 {elem_label} 】"
                ws[f'A{row}'].font      = Font(name=FONT, bold=True, size=11,
                                               color=C['mid_blue'])
                ws[f'A{row}'].alignment = Alignment(horizontal='left')
                ws.row_dimensions[row].height = 18
                row += 1
                months    = sorted(monthly_data.keys())
                stat_keys = list(list(monthly_data.values())[0].keys())
                _hc(ws, row, 1, '통계항목', bg=C['mid_blue'], sz=10)
                ws.column_dimensions['A'].width = 14
                for j, month in enumerate(months):
                    _hc(ws, row, j+2, month, bg=C['mid_blue'], sz=9)
                    ws.column_dimensions[get_column_letter(j+2)].width = 11
                ws.row_dimensions[row].height = 18
                row += 1
                for k, stat_key in enumerate(stat_keys):
                    bg = C['light_blue'] if k % 2 == 0 else C['white']
                    _dc(ws, row, 1, stat_key, bg=bg, bold=True, sz=10, align='left')
                    for j, month in enumerate(months):
                        val = monthly_data.get(month, {}).get(stat_key, '-')
                        _dc(ws, row, j+2, val, bg=bg, sz=10)
                    ws.row_dimensions[row].height = 16
                    row += 1
                row += 2
            row += 2

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 3. 원본 데이터  (요구사항 5: 연도·월 열 추가, Excel Table 생성)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_raw(self, wb, df: pd.DataFrame, selected_cols: list):
        ws = wb.create_sheet("📊 원본 데이터")

        # ── 컬럼 레이아웃 결정 ──
        # A: 날짜, B: 연도, C: 월, [D: 관측소명], D/E+: 기상요소
        j = 1
        self._col['date']  = 'A'; j += 1
        self._col['year']  = 'B'; j += 1
        self._col['month'] = 'C'; j += 1
        if self._has_stn:
            self._col['station'] = 'D'; j += 1

        avail = [c for c in selected_cols if c in df.columns]
        for ck in avail:
            self._col[ck] = get_column_letter(j)
            j += 1

        total_cols = j - 1

        # ── 헤더 ──
        headers = ['날짜', '연도', '월']
        if self._has_stn:
            headers.append('관측소명')
        rename = {'date':'날짜', 'station_name':'관측소명'}
        rename.update(ELEMENT_LABELS)
        for ck in avail:
            headers.append(ELEMENT_LABELS.get(ck, ck))

        for jj, h in enumerate(headers, 1):
            _hc(ws, 1, jj, h, bg=C['dark_blue'], sz=10)
            ws.column_dimensions[get_column_letter(jj)].width = 13
        ws.row_dimensions[1].height = 20

        # ── 데이터 행 ──
        sub = df.copy()
        sub['_date'] = sub['date'].dt.strftime('%Y-%m-%d')

        for i, (_, row_data) in enumerate(sub.iterrows(), start=2):
            bg = C['light_blue'] if i % 2 == 0 else C['white']
            r  = i

            # 날짜 (A)
            _dc(ws, r, 1, row_data['_date'], bg=bg, sz=9)
            # 연도 (B) — 수식 (요구사항 4)
            _dc(ws, r, 2, f'=YEAR(A{r})', bg=bg, sz=9, nf='0')
            # 월 (C) — 수식 (요구사항 4)
            _dc(ws, r, 3, f'=MONTH(A{r})', bg=bg, sz=9, nf='0')

            col_idx = 4
            if self._has_stn:
                _dc(ws, r, col_idx, row_data.get('station_name', ''), bg=bg, sz=9)
                col_idx += 1

            for ck in avail:
                val = row_data.get(ck, None)
                if isinstance(val, float) and np.isnan(val):
                    val = None
                _dc(ws, r, col_idx, val, bg=bg, sz=9,
                    nf='#,##0.0' if isinstance(val, float) else None)
                col_idx += 1

            ws.row_dimensions[r].height = 14

        # ── Excel Table 생성 (PivotTable 소스용) ──
        last_row  = len(sub) + 1
        last_col  = get_column_letter(total_cols)
        tbl_range = f"A1:{last_col}{last_row}"

        tbl = Table(displayName="기상데이터", ref=tbl_range)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True,   showColumnStripes=False
        )
        ws.add_table(tbl)

        # Table이 autoFilter를 자체 포함 — 시트 레벨 별도 설정 불필요
        ws.freeze_panes = 'A2'

        # 나중에 PivotTable XML에서 사용
        self._raw_table_range = tbl_range
        self._raw_col_headers = headers

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 4. 기상 차트 (기존 유지)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_charts(self, wb, df, selected_cols):
        ws = wb.create_sheet("📈 기상 차트")
        ws.sheet_view.showGridLines = False
        df2 = df.copy()
        df2['ym'] = df2['date'].dt.to_period('M').astype(str)
        chart_anchor_row = 1
        dc = 30  # 보조 데이터 시작 컬럼

        def wdata(ws, cs, rs, months, labels, series):
            for k, lbl in enumerate(labels):
                ws.cell(row=rs, column=cs+k, value=lbl)
            for i, mo in enumerate(months):
                ws.cell(row=rs+1+i, column=cs, value=mo)
                for k, s in enumerate(series):
                    ws.cell(row=rs+1+i, column=cs+1+k, value=s.get(mo))
            n    = len(months)
            cats = Reference(ws, min_col=cs, min_row=rs+1, max_row=rs+n)
            refs = [Reference(ws, min_col=cs+1+k, min_row=rs, max_row=rs+n)
                    for k in range(len(series))]
            return cats, refs

        tc = [c for c in ['temp_avg','temp_max','temp_min'] if c in df.columns]
        if tc and any(c in selected_cols for c in tc):
            mt = {c: df2.groupby('ym')[c].mean().round(1).to_dict() for c in tc}
            mos = sorted(set().union(*[d.keys() for d in mt.values()]))
            cats, refs = wdata(ws, dc, 1, mos,
                               ['월']+[ELEMENT_LABELS[c] for c in tc],
                               [mt[c] for c in tc])
            ch = LineChart()
            ch.title = "월별 기온 현황"; ch.style = 10
            ch.y_axis.title = "기온 (°C)"; ch.width, ch.height = 22, 13
            for r in refs: ch.add_data(r, titles_from_data=True)
            ch.set_categories(cats)
            _set_chart_layout(ch)
            ws.add_chart(ch, f'A{chart_anchor_row}')
            chart_anchor_row += 24; dc += len(tc) + 2

        if 'precipitation' in df.columns and 'precipitation' in selected_cols:
            mp = df2.groupby('ym')['precipitation'].sum().round(1).to_dict()
            mos = sorted(mp.keys())
            cats, refs = wdata(ws, dc, chart_anchor_row, mos, ['월','강수량(mm)'], [mp])
            ch2 = BarChart()
            ch2.type='col'; ch2.title="월별 강수량"; ch2.style=10
            ch2.y_axis.title="강수량 (mm)"; ch2.width, ch2.height = 22, 13
            ch2.add_data(refs[0], titles_from_data=True); ch2.set_categories(cats)
            _set_chart_layout(ch2)
            ws.add_chart(ch2, f'A{chart_anchor_row}')
            chart_anchor_row += 24; dc += 3

        if 'humidity' in df.columns and 'humidity' in selected_cols:
            mh = df2.groupby('ym')['humidity'].mean().round(1).to_dict()
            mos = sorted(mh.keys())
            cats, refs = wdata(ws, dc, chart_anchor_row, mos, ['월','평균습도(%)'], [mh])
            ch3 = LineChart()
            ch3.title="월별 평균 습도"; ch3.style=10
            ch3.y_axis.title="습도 (%)"; ch3.width, ch3.height = 22, 13
            ch3.add_data(refs[0], titles_from_data=True); ch3.set_categories(cats)
            _set_chart_layout(ch3)
            ws.add_chart(ch3, f'A{chart_anchor_row}')
            chart_anchor_row += 24; dc += 3

        if 'wind_speed' in df.columns and 'wind_speed' in selected_cols:
            mw = df2.groupby('ym')['wind_speed'].mean().round(1).to_dict()
            mos = sorted(mw.keys())
            cats, refs = wdata(ws, dc, chart_anchor_row, mos, ['월','평균풍속(m/s)'], [mw])
            ch4 = BarChart()
            ch4.type='col'; ch4.title="월별 평균 풍속"; ch4.style=10
            ch4.y_axis.title="풍속 (m/s)"; ch4.width, ch4.height = 22, 13
            ch4.add_data(refs[0], titles_from_data=True); ch4.set_categories(cats)
            _set_chart_layout(ch4)
            ws.add_chart(ch4, f'A{chart_anchor_row}')
            chart_anchor_row += 24; dc += 3

        if 'sunshine' in df.columns and 'sunshine' in selected_cols:
            ms = df2.groupby('ym')['sunshine'].sum().round(1).to_dict()
            mos = sorted(ms.keys())
            cats, refs = wdata(ws, dc, chart_anchor_row, mos, ['월','일조시간(hr)'], [ms])
            ch5 = BarChart()
            ch5.type='col'; ch5.title="월별 일조시간"; ch5.style=10
            ch5.y_axis.title="일조시간 (hr)"; ch5.width, ch5.height = 22, 13
            ch5.add_data(refs[0], titles_from_data=True); ch5.set_categories(cats)
            _set_chart_layout(ch5)
            ws.add_chart(ch5, f'A{chart_anchor_row}')

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 5. 피벗 분석  (요구사항 3·4: ROUND 수식 기반)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_pivot(self, wb, df, selected_cols):
        ws = wb.create_sheet("🔄 피벗 분석")
        ws.sheet_view.showGridLines = True
        ws.freeze_panes = 'C3'
        years = self._years
        avail = [c for c in selected_cols
                 if c in df.columns and c in ELEMENT_LABELS and c not in MODE_ELEMENTS]
        yc  = self._col.get('year',  'B')
        mc  = self._col.get('month', 'C')
        row = 1
        _title(ws, row, 1, 16, '피벗 분석  —  연도 × 월별 기상요소 집계 (수식 기반)', h=26)
        row += 1
        _note(ws, row, 1, 16,
              f'※ {RAW_SHEET} 시트의 데이터를 AVERAGEIFS·SUMIFS 수식으로 집계. '
              '값 수정 없이 원본 데이터 변경 시 자동 갱신됩니다.')
        row += 2

        for ck in avail:
            lbl    = ELEMENT_LABELS[ck]
            is_sum = ck in SUM_ELEMENTS
            fn     = 'SUMIFS' if is_sum else 'AVERAGEIFS'
            agg_l  = '합계' if is_sum else '평균'
            ec     = self._col.get(ck)
            if not ec:
                continue

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
            c = ws.cell(row=row, column=1, value=f"【 {lbl} 】  월별 {agg_l}")
            c.font      = Font(name=FONT, bold=True, size=11, color=C['white'])
            c.fill      = PatternFill(start_color=C['mid_blue'], fill_type='solid')
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 20
            row += 1

            _hc(ws, row, 1, '연도', bg=C['dark_blue'], sz=10)
            for m in range(1, 13):
                _hc(ws, row, m+1, f'{m}월', bg=C['dark_blue'], sz=9)
            _hc(ws, row, 14, f'연간{agg_l}', bg=C['orange'], sz=10)
            ws.row_dimensions[row].height = 18
            row += 1

            data_start = row
            for i, yr in enumerate(years):
                bg = C['light_blue'] if i % 2 == 0 else C['white']
                _dc(ws, row, 1, yr, bg=bg, bold=True, nf='0')
                for m in range(1, 13):
                    cl = get_column_letter(m+1)
                    # 요구사항 3: ROUND 적용, 요구사항 4: 수식 사용
                    f = (f'=ROUND({fn}({RAW_SHEET}!{ec}:{ec},'
                         f'{RAW_SHEET}!{yc}:{yc},{yr},'
                         f'{RAW_SHEET}!{mc}:{mc},{m}),1)')
                    _dc(ws, row, m+1, f, bg=bg,
                        nf='#,##0.0' if is_sum else '0.0')
                # 연간 합계/평균 (이미 ROUND된 셀들의 합산)
                b_l = get_column_letter(2)
                m_l = get_column_letter(13)
                annual_fn = 'SUM' if is_sum else 'AVERAGE'
                _dc(ws, row, 14,
                    f'=ROUND({annual_fn}({b_l}{row}:{m_l}{row}),1)',
                    bg=C['light_orange'], bold=True,
                    nf='#,##0.0' if is_sum else '0.0')
                ws.row_dimensions[row].height = 15
                row += 1

            data_end = row - 1
            for s_lbl, s_fn, s_bg, s_hbg, s_fg in [
                ('전체평균','AVERAGE',C['light_yellow'],C['yellow'],   C['dark_blue']),
                ('최대',    'MAX',    C['light_green'], C['mid_green'],C['white']),
                ('최소',    'MIN',    C['light_blue'],  C['mid_blue'], C['white']),
            ]:
                _hc(ws, row, 1, s_lbl, bg=s_hbg, fg=s_fg, sz=9)
                for m in range(1, 13):
                    cl = get_column_letter(m+1)
                    _dc(ws, row, m+1,
                        f'=ROUND({s_fn}({cl}{data_start}:{cl}{data_end}),1)',
                        bg=s_bg, bold=True,
                        nf='#,##0.0' if is_sum else '0.0')
                n_l = get_column_letter(14)
                _dc(ws, row, 14,
                    f'=ROUND({s_fn}({n_l}{data_start}:{n_l}{data_end}),1)',
                    bg=s_bg, bold=True,
                    nf='#,##0.0' if is_sum else '0.0')
                ws.row_dimensions[row].height = 15
                row += 1
            row += 2

        ws.column_dimensions['A'].width = 8
        for j in range(2, 14):
            ws.column_dimensions[get_column_letter(j)].width = 8
        ws.column_dimensions[get_column_letter(14)].width = 12

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 6. 강수량 분석
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_precipitation(self, wb, df, selected_cols):
        if 'precipitation' not in df.columns or 'precipitation' not in selected_cols:
            return
        ws = wb.create_sheet("🌧️ 강수량 분석")
        ws.sheet_view.showGridLines = False

        pc  = self._col.get('precipitation')
        yc  = self._col.get('year',  'B')
        mc  = self._col.get('month', 'C')
        tc  = self._col.get('temp_avg')
        years = self._years
        n_y   = len(years)

        df2 = df.copy()
        df2['_y'] = df2['date'].dt.year
        df2['_m'] = df2['date'].dt.month

        ROW = 1
        _title(ws, ROW, 1, n_y+3, '강수량 분석', h=28)
        ROW += 1
        _note(ws, ROW, 1, n_y+3,
              f'※ 분석 기간: {df2["date"].min().strftime("%Y-%m-%d")} ~ '
              f'{df2["date"].max().strftime("%Y-%m-%d")}')
        ROW += 2

        _title(ws, ROW, 1, n_y+3, '【표 1】 연도별 기상개황 (강수량)',
               h=20, bg=C['mid_blue'])
        ROW += 1

        _hc(ws, ROW, 1, '구분', bg=C['dark_blue'], sz=10)
        for j, yr in enumerate(years, 2):
            _hc(ws, ROW, j, str(yr), bg=C['dark_blue'], sz=9)
        _hc(ws, ROW, n_y+2, '평균', bg=C['orange'], sz=10)
        ws.row_dimensions[ROW].height = 18
        ROW += 1

        # 연평균기온 (수식 기반)
        t1_rows = []
        if tc:
            t1_rows.append(('연평균기온(℃)', tc, 'AVERAGEIFS', '0.0'))
        t1_rows += [
            ('총강수량(mm)',     pc, 'SUMIFS',     '#,##0.0'),
            ('월평균강수량(mm)', pc, 'AVG_MONTHLY', '#,##0.0'),
        ]
        data_rows_t1 = []
        for i, (lbl, col_k, fn_type, nf) in enumerate(t1_rows):
            bg = C['light_blue'] if i % 2 == 0 else C['white']
            _dc(ws, ROW, 1, lbl, bg=bg, bold=True, align='left')
            data_rows_t1.append(ROW)
            for j, yr in enumerate(years, 2):
                if fn_type == 'AVERAGEIFS':
                    f = f'=ROUND(AVERAGEIFS({RAW_SHEET}!{col_k}:{col_k},{RAW_SHEET}!{yc}:{yc},{yr}),1)'
                elif fn_type == 'SUMIFS':
                    f = f'=ROUND(SUMIFS({RAW_SHEET}!{col_k}:{col_k},{RAW_SHEET}!{yc}:{yc},{yr}),1)'
                else:  # AVG_MONTHLY
                    f = f'=ROUND(SUMIFS({RAW_SHEET}!{col_k}:{col_k},{RAW_SHEET}!{yc}:{yc},{yr})/12,1)'
                _dc(ws, ROW, j, f, bg=bg, nf=nf)
            b_l = get_column_letter(2)
            e_l = get_column_letter(n_y+1)
            _dc(ws, ROW, n_y+2,
                f'=ROUND(AVERAGE({b_l}{ROW}:{e_l}{ROW}),1)',
                bg=C['light_orange'], bold=True, nf=nf)
            ws.row_dimensions[ROW].height = 16
            ROW += 1

        ROW += 3

        # ── 표 2: 연도×월 강수량 상세 (수식 기반) ──
        _title(ws, ROW, 1, 15, '【표 2】 연도별 월별 강수량 (mm)', h=20, bg=C['mid_blue'])
        ROW += 1
        _hc(ws, ROW, 1, '연도', bg=C['dark_blue'], sz=10)
        _hc(ws, ROW, 2, '연계', bg=C['dark_blue'], sz=10)
        for m in range(1, 13):
            _hc(ws, ROW, m+2, f'{m}월', bg=C['dark_blue'], sz=9)
        ws.row_dimensions[ROW].height = 18
        ROW += 1

        t2_start = ROW
        for i, yr in enumerate(years):
            bg = C['light_blue'] if i % 2 == 0 else C['white']
            _dc(ws, ROW, 1, yr, bg=bg, bold=True, nf='0')
            # 연계 = sum of monthly cells in this row (ROUND 적용 후 합산)
            m_start = get_column_letter(3)
            m_end   = get_column_letter(14)
            _dc(ws, ROW, 2, f'=ROUND(SUM({m_start}{ROW}:{m_end}{ROW}),1)',
                bg=bg, bold=True, nf='#,##0.0')
            for m in range(1, 13):
                f = (f'=ROUND(SUMIFS({RAW_SHEET}!{pc}:{pc},'
                     f'{RAW_SHEET}!{yc}:{yc},{yr},'
                     f'{RAW_SHEET}!{mc}:{mc},{m}),1)')
                _dc(ws, ROW, m+2, f, bg=bg, nf='#,##0.0')
            ws.row_dimensions[ROW].height = 15
            ROW += 1

        t2_end = ROW - 1
        for s_lbl, s_fn, s_bg, s_hbg, s_fg in [
            ('평균','AVERAGE',C['light_yellow'],C['yellow'],   C['dark_blue']),
            ('최대','MAX',    C['light_green'], C['mid_green'],C['white']),
            ('최소','MIN',    C['light_blue'],  C['mid_blue'], C['white']),
        ]:
            _hc(ws, ROW, 1, s_lbl, bg=s_hbg, fg=s_fg, sz=9)
            for j in range(2, 15):
                cl = get_column_letter(j)
                _dc(ws, ROW, j,
                    f'=ROUND({s_fn}({cl}{t2_start}:{cl}{t2_end}),1)',
                    bg=s_bg, bold=True, nf='#,##0.0')
            ws.row_dimensions[ROW].height = 15
            ROW += 1

        ROW += 3

        # ── 표 3: 월별 통계 (수식 기반) ──
        _title(ws, ROW, 1, 15, '【표 3】 월별 강수량 통계 (mm)', h=20, bg=C['mid_blue'])
        ROW += 1
        _hc(ws, ROW, 1, '구분', bg=C['dark_blue'], sz=10)
        for m in range(1, 13):
            _hc(ws, ROW, m+1, f'{m}월', bg=C['dark_blue'], sz=9)
        _hc(ws, ROW, 14, '연평균', bg=C['orange'], sz=10)
        ws.row_dimensions[ROW].height = 18
        ROW += 1

        t3_rows = [
            ('월 평균(mm)', 'AVERAGEIFS', C['light_blue']),
            ('최대(mm)',    'MAX_IF',     C['light_green']),
            ('최소(mm)',    'MIN_IF',     C['light_yellow']),
        ]
        for lbl, fn_t, bg in t3_rows:
            _dc(ws, ROW, 1, lbl, bg=bg, bold=True, align='left')
            for m in range(1, 13):
                if fn_t == 'AVERAGEIFS':
                    f = (f'=ROUND(AVERAGEIFS({RAW_SHEET}!{pc}:{pc},'
                         f'{RAW_SHEET}!{mc}:{mc},{m}),1)')
                elif fn_t == 'MAX_IF':
                    # MAX(IF()) 배열수식 — MAXIFS 대신 사용 (Excel의 @ 자동추가 방지)
                    f = (f'=ROUND(MAX(IF({RAW_SHEET}!{mc}:{mc}={m},'
                         f'{RAW_SHEET}!{pc}:{pc})),1)')
                else:
                    # MIN(IF()) 배열수식
                    f = (f'=ROUND(MIN(IF({RAW_SHEET}!{mc}:{mc}={m},'
                         f'{RAW_SHEET}!{pc}:{pc})),1)')
                _dc(ws, ROW, m+1, f, bg=bg, nf='#,##0.0')
            b_l = get_column_letter(2)
            m_l = get_column_letter(13)
            _dc(ws, ROW, 14,
                f'=ROUND(AVERAGE({b_l}{ROW}:{m_l}{ROW}),1)',
                bg=C['light_orange'], bold=True, nf='#,##0.0')
            ws.row_dimensions[ROW].height = 16
            ROW += 1

        ROW += 3

        # ── 차트용 보조 데이터 ──
        AUX   = max(17, n_y + 4)
        AUX_M = AUX + 4
        ws.cell(row=1, column=AUX,   value='연도')
        ws.cell(row=1, column=AUX+1, value='연강수량(mm)')
        ws.cell(row=1, column=AUX+2, value='월평균(mm)')
        for i, yr in enumerate(years):
            ws.cell(row=2+i, column=AUX,   value=yr)
            ws.cell(row=2+i, column=AUX+1,
                    value=round(float(df2[df2['_y']==yr]['precipitation'].sum()), 1))
            ws.cell(row=2+i, column=AUX+2,
                    value=round(float(df2[df2['_y']==yr]['precipitation'].sum())/12, 1))

        ws.cell(row=1, column=AUX_M,   value='월')
        ws.cell(row=1, column=AUX_M+1, value='월평균(mm)')
        ms = df2.groupby('_m')['precipitation'].agg(['mean','max','min']).round(1)
        for m in range(1, 13):
            ws.cell(row=1+m, column=AUX_M,   value=f'{m}월')
            ws.cell(row=1+m, column=AUX_M+1,
                    value=float(ms.loc[m,'mean']) if m in ms.index else 0)

        for ci in range(AUX, AUX+10):
            ws.column_dimensions[get_column_letter(ci)].width = 8

        # ── 차트 ──
        ch_a = BarChart()
        ch_a.type='col'; ch_a.title="연도별 강수량 현황"; ch_a.style=10
        ch_a.y_axis.title="연강수량 (mm)"; ch_a.width, ch_a.height = 24, 14
        d_bar = Reference(ws, min_col=AUX+1, max_col=AUX+1, min_row=1, max_row=1+n_y)
        c_bar = Reference(ws, min_col=AUX, min_row=2, max_row=1+n_y)
        ch_a.add_data(d_bar, titles_from_data=True)
        ch_a.set_categories(c_bar)
        ch_a.series[0].graphicalProperties.solidFill = '2E75B6'
        ln_a = LineChart()
        d_ln = Reference(ws, min_col=AUX+2, max_col=AUX+2, min_row=1, max_row=1+n_y)
        ln_a.add_data(d_ln, titles_from_data=True)
        ln_a.set_categories(c_bar)
        ln_a.y_axis.axId=200; ln_a.y_axis.title="월평균(mm)"
        ln_a.series[0].graphicalProperties.line.solidFill='C55A11'
        ln_a.series[0].graphicalProperties.line.width=20000
        ln_a.series[0].marker=Marker(symbol='circle', size=6)
        ch_a += ln_a
        _set_chart_layout(ch_a)
        ws.add_chart(ch_a, f'A{ROW}')

        ch_b = BarChart()
        ch_b.type='col'; ch_b.title="월별 평균 강수량"; ch_b.style=10
        ch_b.y_axis.title="강수량 (mm)"; ch_b.width, ch_b.height = 24, 14
        d_mo = Reference(ws, min_col=AUX_M+1, max_col=AUX_M+1, min_row=1, max_row=13)
        c_mo = Reference(ws, min_col=AUX_M, min_row=2, max_row=13)
        ch_b.add_data(d_mo, titles_from_data=True)
        ch_b.set_categories(c_mo)
        ch_b.series[0].graphicalProperties.solidFill='4472C4'
        _set_chart_layout(ch_b)
        ws.add_chart(ch_b, f'M{ROW}')

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        for j in range(3, 15):
            ws.column_dimensions[get_column_letter(j)].width = 8

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 7. 기상 특성 검토  (NEW — 요구사항 1)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_weather_chars(self, wb, df, selected_cols):
        ws = wb.create_sheet("🌡️ 기상 특성 검토")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'B5'

        yc = self._col.get('year',  'B')
        mc = self._col.get('month', 'C')

        # 분석 대상 요소
        ELEMS = [
            ('temp_avg',     '평균기온(℃)',   'AVERAGEIFS', '0.0'),
            ('temp_max',     '최고기온(℃)',   'AVERAGEIFS', '0.0'),
            ('temp_min',     '최저기온(℃)',   'AVERAGEIFS', '0.0'),
            ('precipitation','강수량(mm)',     'SUMIFS_AVG', '#,##0.0'),
            ('humidity',     '습도(%)',        'AVERAGEIFS', '0.0'),
            ('wind_speed',   '풍속(m/s)',      'AVERAGEIFS', '0.0'),
            ('sunshine',     '일조시간(hr)',   'SUMIFS_AVG', '#,##0.0'),
        ]
        avail_elems = [(ck, lbl, fn, nf) for ck, lbl, fn, nf in ELEMS
                       if ck in selected_cols and ck in self._col]
        n_elem = len(avail_elems)
        if n_elem == 0:
            return

        # ── 시트 제목 ──
        ROW = 1
        _title(ws, ROW, 1, n_elem+2, '기상 특성 검토', h=28)
        ROW += 1
        _note(ws, ROW, 1, n_elem+2,
              f'※ 전체 관측 기간 기후값  |  강수량·일조시간=연평균 월합계, 기온·습도·풍속=월평균  |  '
              f'수식 기반 — 원본 데이터 변경 시 자동 갱신')
        ROW += 2

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 표 1: 월별 기후 특성 (12개월 + 연간)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        _title(ws, ROW, 1, n_elem+2, '【표 1】 월별 기후 특성', h=20, bg=C['mid_blue'])
        ROW += 1

        _hc(ws, ROW, 1, '월', bg=C['dark_blue'], sz=10)
        for j, (ck, lbl, fn, nf) in enumerate(avail_elems, 2):
            _hc(ws, ROW, j, lbl, bg=C['dark_blue'], sz=9, wrap=True)
        ws.row_dimensions[ROW].height = 28
        ROW += 1

        monthly_data_start = ROW  # 차트용
        monthly_rows = {}  # m → row number (계절 집계에서 참조)

        for m in range(1, 13):
            bg = C['light_blue'] if m % 2 == 0 else C['white']
            _dc(ws, ROW, 1, f'{m}월', bg=bg, bold=True)
            monthly_rows[m] = ROW
            for j, (ck, lbl, fn, nf) in enumerate(avail_elems, 2):
                ec = self._col[ck]
                if fn == 'AVERAGEIFS':
                    f = (f'=ROUND(AVERAGEIFS({RAW_SHEET}!{ec}:{ec},'
                         f'{RAW_SHEET}!{mc}:{mc},{m}),1)')
                else:  # SUMIFS_AVG: 해당 월의 연간 합계를 연수로 나눔
                    n_yr = len(self._years)
                    f = (f'=ROUND(SUMIFS({RAW_SHEET}!{ec}:{ec},'
                         f'{RAW_SHEET}!{mc}:{mc},{m})/{n_yr},1)')
                _dc(ws, ROW, j, f, bg=bg, nf=nf)
            ws.row_dimensions[ROW].height = 15
            ROW += 1

        monthly_data_end = ROW - 1

        # 연간 집계행 (ROUND된 월별 셀들의 합계/평균)
        _hc(ws, ROW, 1, '연간', bg=C['orange'], fg=C['white'], sz=10)
        for j, (ck, lbl, fn, nf) in enumerate(avail_elems, 2):
            cl = get_column_letter(j)
            if fn in ('SUMIFS_AVG',):  # 강수량·일조시간은 연간합계
                f = f'=ROUND(SUM({cl}{monthly_data_start}:{cl}{monthly_data_end}),1)'
            else:
                f = f'=ROUND(AVERAGE({cl}{monthly_data_start}:{cl}{monthly_data_end}),1)'
            _dc(ws, ROW, j, f, bg=C['light_orange'], bold=True, nf=nf)
        ws.row_dimensions[ROW].height = 18
        annual_row = ROW
        ROW += 3

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 표 2: 계절별 기상 특성
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        _title(ws, ROW, 1, n_elem+2, '【표 2】 계절별 기상 특성', h=20, bg=C['mid_blue'])
        ROW += 1
        _hc(ws, ROW, 1, '계절', bg=C['dark_blue'], sz=10)
        for j, (ck, lbl, fn, nf) in enumerate(avail_elems, 2):
            _hc(ws, ROW, j, lbl, bg=C['dark_blue'], sz=9, wrap=True)
        ws.row_dimensions[ROW].height = 28
        ROW += 1

        SEASONS = [
            ('봄(3~5월)',    [3,4,5]),
            ('여름(6~8월)', [6,7,8]),
            ('가을(9~11월)',[9,10,11]),
            ('겨울(12~2월)',[12,1,2]),
        ]
        season_bgs = [C['light_green'], 'FCE4D6', 'FFF2CC', C['light_blue']]

        for si, ((s_lbl, s_months), s_bg) in enumerate(zip(SEASONS, season_bgs)):
            _dc(ws, ROW, 1, s_lbl, bg=s_bg, bold=True)
            for j, (ck, lbl, fn, nf) in enumerate(avail_elems, 2):
                cl = get_column_letter(j)
                # 해당 월들의 표 1 데이터를 참조 (ROUND된 값들의 평균/합)
                month_refs = ','.join(f'{cl}{monthly_rows[m]}' for m in s_months)
                if fn in ('SUMIFS_AVG',):
                    f = f'=ROUND(SUM({month_refs}),1)'
                else:
                    f = f'=ROUND(AVERAGE({month_refs}),1)'
                _dc(ws, ROW, j, f, bg=s_bg, nf=nf)
            ws.row_dimensions[ROW].height = 16
            ROW += 1

        ROW += 3

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 표 3: 극값 분석 (수식 기반)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        _title(ws, ROW, 1, n_elem+2, '【표 3】 극값 분석 (관측 기간 전체)', h=20, bg=C['mid_blue'])
        ROW += 1
        _hc(ws, ROW, 1, '구분', bg=C['dark_blue'], sz=10)
        for j, (ck, lbl, fn, nf) in enumerate(avail_elems, 2):
            _hc(ws, ROW, j, lbl, bg=C['dark_blue'], sz=9, wrap=True)
        ws.row_dimensions[ROW].height = 28
        ROW += 1

        for s_lbl, s_fn, bg in [
            ('최대값', 'MAX', C['light_orange']),
            ('최솟값', 'MIN', C['light_blue']),
        ]:
            _dc(ws, ROW, 1, s_lbl, bg=bg, bold=True)
            for j, (ck, lbl, fn, nf) in enumerate(avail_elems, 2):
                ec = self._col[ck]
                f = f'=ROUND({s_fn}({RAW_SHEET}!{ec}:{ec}),1)'
                _dc(ws, ROW, j, f, bg=bg, bold=True, nf=nf)
            ws.row_dimensions[ROW].height = 16
            ROW += 1

        ROW += 3

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 차트 1: 월별 기온 (꺾은선)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        temp_elems = [(j, ck) for j, (ck,_,_,_) in enumerate(avail_elems, 2)
                      if ck in ('temp_avg','temp_max','temp_min')]
        if temp_elems:
            ch_temp = LineChart()
            ch_temp.title = "월별 기온 특성 (℃)"; ch_temp.style = 10
            ch_temp.y_axis.title = "기온 (℃)"; ch_temp.x_axis.title = "월"
            ch_temp.width, ch_temp.height = 22, 13
            cats = Reference(ws, min_col=1, min_row=monthly_data_start,
                             max_row=monthly_data_end)
            for j, _ in temp_elems:
                d = Reference(ws, min_col=j, min_row=ROW-ROW+monthly_data_start-1,
                              max_row=monthly_data_end)
                # Re-reference correctly
            # Simple approach: reference the columns directly
            for j, ck in temp_elems:
                d_ref = Reference(ws, min_col=j, min_row=monthly_data_start-1,
                                  max_row=monthly_data_end)
                ch_temp.add_data(d_ref, titles_from_data=True)
            ch_temp.set_categories(cats)
            TEMP_C = ['4472C4','ED7D31','A9D18E']
            for k, s in enumerate(ch_temp.series):
                s.graphicalProperties.line.solidFill = TEMP_C[k % 3]
                s.graphicalProperties.line.width     = 20000
                s.marker = Marker(symbol='circle', size=5)
            _set_chart_layout(ch_temp)
            ws.add_chart(ch_temp, f'A{ROW}')

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 차트 2: 월별 강수량 (막대)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        precip_j = next((j for j,(ck,_,_,_) in enumerate(avail_elems, 2)
                         if ck == 'precipitation'), None)
        if precip_j:
            ch_p = BarChart()
            ch_p.type='col'; ch_p.title="월별 평균 강수량 (mm)"; ch_p.style=10
            ch_p.y_axis.title="강수량 (mm)"; ch_p.x_axis.title="월"
            ch_p.width, ch_p.height = 22, 13
            d_p = Reference(ws, min_col=precip_j,
                            min_row=monthly_data_start-1, max_row=monthly_data_end)
            ch_p.add_data(d_p, titles_from_data=True)
            ch_p.set_categories(cats)
            ch_p.series[0].graphicalProperties.solidFill = '2E75B6'
            _set_chart_layout(ch_p)
            ws.add_chart(ch_p, f'L{ROW}')

        # 열 너비
        ws.column_dimensions['A'].width = 14
        for j in range(2, n_elem+2):
            ws.column_dimensions[get_column_letter(j)].width = 13

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 8. 기후변화 검토  (NEW — 요구사항 2)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_climate_change(self, wb, df, selected_cols):
        ws = wb.create_sheet("🌍 기후변화 검토")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'B5'

        yc     = self._col.get('year',  'B')
        tc     = self._col.get('temp_avg')
        pc     = self._col.get('precipitation')
        years  = self._years
        n_year = len(years)
        if n_year < 2:
            return

        ROW = 1
        _title(ws, ROW, 1, 10, '기후변화 검토', h=28)
        ROW += 1
        _note(ws, ROW, 1, 10,
              f'※ 분석 기간: {years[0]}~{years[-1]}년 ({n_year}년)  '
              '|  수식 기반 — 원본 데이터 변경 시 자동 갱신')
        ROW += 2

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 표 1: 연도별 기상 통계
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        _title(ws, ROW, 1, 10, '【표 1】 연도별 기상 통계', h=20, bg=C['mid_blue'])
        ROW += 1

        headers = ['연도']
        if tc: headers += ['연평균기온(℃)', '최고기온(℃)', '최저기온(℃)']
        if pc: headers += ['총강수량(mm)', '월평균강수량(mm)']
        if 'humidity' in self._col and 'humidity' in selected_cols:
            headers.append('평균습도(%)')

        for j, h in enumerate(headers, 1):
            _hc(ws, ROW, j, h, bg=C['dark_blue'], sz=9, wrap=True)
            ws.column_dimensions[get_column_letter(j)].width = 14
        ws.row_dimensions[ROW].height = 28
        ROW += 1

        t1_data_start = ROW
        yr_rows = {}  # year → row (이동평균 수식 참조용)

        for i, yr in enumerate(years):
            bg = C['light_blue'] if i % 2 == 0 else C['white']
            col = 1
            _dc(ws, ROW, col, yr, bg=bg, bold=True, nf='0'); col += 1
            if tc:
                for sub_fn in ('AVERAGEIFS', 'MAX_IF', 'MIN_IF'):
                    if sub_fn == 'AVERAGEIFS':
                        f = (f'=ROUND(AVERAGEIFS({RAW_SHEET}!{tc}:{tc},'
                             f'{RAW_SHEET}!{yc}:{yc},{yr}),1)')
                    elif sub_fn == 'MAX_IF':
                        f = (f'=ROUND(MAX(IF({RAW_SHEET}!{yc}:{yc}={yr},'
                             f'{RAW_SHEET}!{tc}:{tc})),1)')
                    else:
                        f = (f'=ROUND(MIN(IF({RAW_SHEET}!{yc}:{yc}={yr},'
                             f'{RAW_SHEET}!{tc}:{tc})),1)')
                    _dc(ws, ROW, col, f, bg=bg, nf='0.0'); col += 1
            if pc:
                _dc(ws, ROW, col,
                    f'=ROUND(SUMIFS({RAW_SHEET}!{pc}:{pc},{RAW_SHEET}!{yc}:{yc},{yr}),1)',
                    bg=bg, nf='#,##0.0'); col += 1
                _dc(ws, ROW, col,
                    f'=ROUND(SUMIFS({RAW_SHEET}!{pc}:{pc},{RAW_SHEET}!{yc}:{yc},{yr})/12,1)',
                    bg=bg, nf='#,##0.0'); col += 1
            if 'humidity' in self._col and 'humidity' in selected_cols:
                hc2 = self._col['humidity']
                _dc(ws, ROW, col,
                    f'=ROUND(AVERAGEIFS({RAW_SHEET}!{hc2}:{hc2},{RAW_SHEET}!{yc}:{yc},{yr}),1)',
                    bg=bg, nf='0.0'); col += 1
            ws.row_dimensions[ROW].height = 15
            yr_rows[yr] = ROW
            ROW += 1

        t1_data_end = ROW - 1

        # 집계행 (평균/최대/최소)
        for s_lbl, s_fn, s_bg, s_hbg in [
            ('전체평균','AVERAGE',C['light_yellow'],C['yellow']),
            ('최대',    'MAX',    C['light_green'], C['mid_green']),
            ('최소',    'MIN',    C['light_blue'],  C['mid_blue']),
        ]:
            _hc(ws, ROW, 1, s_lbl, bg=s_hbg,
                fg=C['dark_blue'] if s_hbg==C['yellow'] else C['white'], sz=9)
            for j in range(2, len(headers)+1):
                cl = get_column_letter(j)
                _dc(ws, ROW, j,
                    f'=ROUND({s_fn}({cl}{t1_data_start}:{cl}{t1_data_end}),1)',
                    bg=s_bg, bold=True, nf='0.0')
            ws.row_dimensions[ROW].height = 15
            ROW += 1

        ROW += 3

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 표 2: 기간별 비교 (전반기 vs 후반기)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        _title(ws, ROW, 1, 10, '【표 2】 전반기·후반기 기후 비교', h=20, bg=C['mid_blue'])
        ROW += 1

        mid = n_year // 2
        y1_s, y1_e = years[0],   years[mid-1]
        y2_s, y2_e = years[mid], years[-1]

        for j, h in enumerate(['구분', '기간', '연수(년)'] + [h for h in headers[1:]], 1):
            _hc(ws, ROW, j, h, bg=C['dark_blue'], sz=9, wrap=True)
        ws.row_dimensions[ROW].height = 28
        ROW += 1

        for period_lbl, yr_start, yr_end in [
            ('전반기', y1_s, y1_e),
            ('후반기', y2_s, y2_e),
        ]:
            bg = C['light_blue'] if period_lbl == '전반기' else C['light_orange']
            _dc(ws, ROW, 1, period_lbl, bg=bg, bold=True)
            _dc(ws, ROW, 2, f'{yr_start}~{yr_end}년', bg=bg)
            _dc(ws, ROW, 3, yr_end - yr_start + 1, bg=bg, nf='0')
            col = 4
            if tc:
                # 전반기/후반기 연평균기온들의 평균
                yr_list = [yr for yr in years if yr_start <= yr <= yr_end]
                temp_refs = ','.join(f'B{yr_rows[y]}' for y in yr_list if y in yr_rows)
                if temp_refs:
                    _dc(ws, ROW, col,
                        f'=ROUND(AVERAGE({temp_refs}),1)',
                        bg=bg, nf='0.0'); col += 1
                    # 최고기온 평균
                    mx_refs = ','.join(f'C{yr_rows[y]}' for y in yr_list if y in yr_rows)
                    _dc(ws, ROW, col, f'=ROUND(AVERAGE({mx_refs}),1)', bg=bg, nf='0.0'); col += 1
                    mn_refs = ','.join(f'D{yr_rows[y]}' for y in yr_list if y in yr_rows)
                    _dc(ws, ROW, col, f'=ROUND(AVERAGE({mn_refs}),1)', bg=bg, nf='0.0'); col += 1
            if pc:
                precip_col_letter = get_column_letter(
                    (4 if not tc else 4+3) if tc else 4)
                # 강수량 컬럼 인덱스 계산
                p_col_idx = 4 + (3 if tc else 0)
                yr_list2 = [yr for yr in years if yr_start <= yr <= yr_end]
                p_refs = ','.join(f'{get_column_letter(p_col_idx)}{yr_rows[y]}'
                                   for y in yr_list2 if y in yr_rows)
                if p_refs:
                    _dc(ws, ROW, col, f'=ROUND(AVERAGE({p_refs}),1)', bg=bg, nf='#,##0.0')
                    col += 1
                    pm_refs = ','.join(f'{get_column_letter(p_col_idx+1)}{yr_rows[y]}'
                                        for y in yr_list2 if y in yr_rows)
                    _dc(ws, ROW, col, f'=ROUND(AVERAGE({pm_refs}),1)', bg=bg, nf='#,##0.0')
                    col += 1
            ws.row_dimensions[ROW].height = 16
            ROW += 1

        ROW += 3

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 표 3: 5년·10년 이동평균
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        if tc and n_year >= 5:
            _title(ws, ROW, 1, 6, '【표 3】 이동평균 분석 (연평균기온)', h=20, bg=C['mid_blue'])
            ROW += 1
            for j, h in enumerate(['연도','연평균기온(℃)','5년이동평균(℃)','10년이동평균(℃)'], 1):
                _hc(ws, ROW, j, h, bg=C['dark_blue'], sz=9)
                ws.column_dimensions[get_column_letter(j)].width = 16
            ws.row_dimensions[ROW].height = 18
            ROW += 1

            t3_data_start = ROW
            for i, yr in enumerate(years):
                bg = C['light_blue'] if i % 2 == 0 else C['white']
                _dc(ws, ROW, 1, yr, bg=bg, bold=True, nf='0')
                # 연평균기온 (표1 참조)
                _dc(ws, ROW, 2, f'=B{yr_rows[yr]}', bg=bg, nf='0.0')
                # 5년 이동평균 (현재 포함 이전 5개 ROUND된 값의 평균)
                if i >= 4:
                    rows_5 = ','.join(f'B{yr_rows[years[i-k]]}' for k in range(5))
                    _dc(ws, ROW, 3, f'=ROUND(AVERAGE({rows_5}),1)', bg=bg, nf='0.0')
                else:
                    _dc(ws, ROW, 3, '', bg=bg)
                # 10년 이동평균
                if i >= 9:
                    rows_10 = ','.join(f'B{yr_rows[years[i-k]]}' for k in range(10))
                    _dc(ws, ROW, 4, f'=ROUND(AVERAGE({rows_10}),1)', bg=bg, nf='0.0')
                else:
                    _dc(ws, ROW, 4, '', bg=bg)
                ws.row_dimensions[ROW].height = 15
                ROW += 1

            t3_data_end = ROW - 1
            ROW += 3

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 차트: 연평균기온 추이 + 이동평균
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        if tc and n_year >= 5:
            ch_t = LineChart()
            ch_t.title = f"연평균기온 변화 추이 ({years[0]}~{years[-1]})"; ch_t.style = 10
            ch_t.y_axis.title = "기온 (℃)"; ch_t.x_axis.title = "연도"
            ch_t.width, ch_t.height = 26, 14

            cats = Reference(ws, min_col=1, min_row=t3_data_start, max_row=t3_data_end)
            for col_idx, (name, color, width) in enumerate([
                ('연평균기온','4472C4',15000),
                ('5년이동평균','ED7D31',25000),
                ('10년이동평균','FF0000',25000),
            ], 2):
                d = Reference(ws, min_col=col_idx,
                              min_row=t3_data_start-1, max_row=t3_data_end)
                ch_t.add_data(d, titles_from_data=True)
            ch_t.set_categories(cats)
            for k, (color, w) in enumerate([('4472C4',15000),('ED7D31',25000),('FF0000',25000)]):
                if k < len(ch_t.series):
                    ch_t.series[k].graphicalProperties.line.solidFill = color
                    ch_t.series[k].graphicalProperties.line.width = w
            _set_chart_layout(ch_t)
            ws.add_chart(ch_t, f'A{ROW}')
            ROW += 30

        # 차트: 연강수량 추이
        if pc and n_year >= 2:
            # 연강수량은 표1의 5번째 열(or pc based)
            p_col_chart_idx = 4 + (3 if tc else 0)  # 표1에서 총강수량 열 인덱스
            ch_p = BarChart()
            ch_p.type='col'; ch_p.title=f"연강수량 변화 추이 ({years[0]}~{years[-1]})"
            ch_p.style=10; ch_p.y_axis.title="강수량 (mm)"
            ch_p.width, ch_p.height = 26, 14

            cats2 = Reference(ws, min_col=1, min_row=t1_data_start, max_row=t1_data_end)
            d_p = Reference(ws, min_col=p_col_chart_idx,
                            min_row=t1_data_start-1, max_row=t1_data_end)
            ch_p.add_data(d_p, titles_from_data=True)
            ch_p.set_categories(cats2)
            ch_p.series[0].graphicalProperties.solidFill = '2E75B6'
            _set_chart_layout(ch_p)
            ws.add_chart(ch_p, f'M{ROW - (30 if (tc and n_year>=5) else 0)}')

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 9. 피벗작업  — Excel PivotTable 자리 마련
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_pivot_work(self, wb, df, selected_cols):
        ws = wb.create_sheet("피벗작업")
        ws.sheet_view.showGridLines = True
        # PivotTable은 A3에 삽입됩니다 (_inject_pivot_table에서 처리)
        _title(ws, 1, 1, 8,
               '피벗 테이블  —  아래 피벗 테이블을 클릭하면 필드 목록이 표시됩니다.', h=24)
        _note(ws, 2, 1, 8,
              '※ Excel 피벗 테이블  |  소스: 원본 데이터(기상데이터 Table)  '
              '|  필드를 드래그해 자유롭게 분석하세요.')
        # A3부터 PivotTable이 삽입됩니다 (inject 단계)
        ws.sheet_view.showGridLines = True

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 10. 누적강수량 분석
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_cumulative_precip(self, wb, df, selected_cols):
        if 'precipitation' not in df.columns or 'precipitation' not in selected_cols:
            return
        ws = wb.create_sheet("누적강수량 분석")
        ws.sheet_view.showGridLines = True
        ws.freeze_panes = 'B2'

        pc    = self._col.get('precipitation')
        yc    = self._col.get('year',  'B')
        mc    = self._col.get('month', 'C')
        years = self._years
        n     = len(years)

        df2 = df.copy()
        df2['_y'] = df2['date'].dt.year
        df2['_m'] = df2['date'].dt.month
        mp = df2.pivot_table(values='precipitation', index='_m',
                             columns='_y', aggfunc='sum').round(1)

        yr2col = {yr: i+2 for i, yr in enumerate(years)}
        col_10 = n+2; col_20 = n+3; col_30 = n+4

        def cl(yr):
            return get_column_letter(yr2col[yr])

        def yr_avg_formula(yr_list, row):
            cols = '+'.join(f'{cl(y)}{row}' for y in yr_list)
            return f'=ROUND(({cols})/{len(yr_list)},1)'

        def avg_range(ys, k):
            return ys[-k:] if len(ys) >= k else ys

        yr10 = avg_range(years, 10)
        yr20 = avg_range(years, 20)
        yr30 = avg_range(years, 30)

        def write_hdr(ws, row, years):
            ws.cell(row=row, column=1).border = _thin()
            ws.column_dimensions['A'].width = 10
            for yr in years:
                _hc(ws, row, yr2col[yr], yr, bg=C['dark_blue'], sz=9)
                ws.column_dimensions[get_column_letter(yr2col[yr])].width = 8
            _hc(ws, row, col_10, '10년평균', bg=C['orange'], sz=9)
            _hc(ws, row, col_20, '20년평균', bg=C['orange'], sz=9)
            _hc(ws, row, col_30, '30년평균', bg=C['orange'], sz=9)
            for c3 in [col_10, col_20, col_30]:
                ws.column_dimensions[get_column_letter(c3)].width = 10
            ws.row_dimensions[row].height = 18

        SUMMER = [('6+7',[6,7]),('6+7+8',[6,7,8]),('7+8',[7,8]),
                  ('7+8+9',[7,8,9]),('6+7+8+9',[6,7,8,9])]

        ROW = 1
        _title(ws, ROW, 1, col_30, '【구간 ①】 연도별 월별 강수량 (mm) — 절대값', h=24)
        ROW += 1
        write_hdr(ws, ROW, years)
        ROW += 1

        M_ROW = {}
        for m in range(1, 13):
            bg = C['light_blue'] if m % 2 == 0 else C['white']
            _dc(ws, ROW, 1, m, bg=bg, bold=True, nf='0')
            for yr in years:
                # 수식 기반 (요구사항 4)
                f = (f'=ROUND(SUMIFS({RAW_SHEET}!{pc}:{pc},'
                     f'{RAW_SHEET}!{yc}:{yc},{yr},'
                     f'{RAW_SHEET}!{mc}:{mc},{m}),1)')
                _dc(ws, ROW, yr2col[yr], f, bg=bg, nf='#,##0.0')
            _dc(ws, ROW, col_10, yr_avg_formula(yr10, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            _dc(ws, ROW, col_20, yr_avg_formula(yr20, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            _dc(ws, ROW, col_30, yr_avg_formula(yr30, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            ws.row_dimensions[ROW].height = 15
            M_ROW[m] = ROW
            ROW += 1

        # 연간합계 — ROUND된 월별 셀들의 합
        ROW_ANN = ROW
        _hc(ws, ROW, 1, '합계', bg=C['mid_blue'], sz=9)
        for yr in years:
            m_refs = '+'.join(f'{cl(yr)}{M_ROW[m]}' for m in range(1,13))
            _dc(ws, ROW, yr2col[yr], f'=ROUND(SUM({m_refs}),1)',
                bg=C['light_green'], bold=True, nf='#,##0.0')
        _dc(ws, ROW, col_10, yr_avg_formula(yr10, ROW), bg=C['light_yellow'], bold=True, nf='#,##0.0')
        _dc(ws, ROW, col_20, yr_avg_formula(yr20, ROW), bg=C['light_yellow'], bold=True, nf='#,##0.0')
        _dc(ws, ROW, col_30, yr_avg_formula(yr30, ROW), bg=C['light_yellow'], bold=True, nf='#,##0.0')
        ws.row_dimensions[ROW].height = 16
        ROW += 2

        COMBO_ROW1 = {}
        for combo_lbl, months in SUMMER:
            _hc(ws, ROW, 1, combo_lbl, bg=C['mid_blue'], fg=C['white'], sz=9)
            for yr in years:
                m_refs = '+'.join(f'{cl(yr)}{M_ROW[m]}' for m in months)
                _dc(ws, ROW, yr2col[yr], f'=ROUND(SUM({m_refs}),1)',
                    bg=C['light_blue'], nf='#,##0.0')
            _dc(ws, ROW, col_10, yr_avg_formula(yr10, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            _dc(ws, ROW, col_20, yr_avg_formula(yr20, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            _dc(ws, ROW, col_30, yr_avg_formula(yr30, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            ws.row_dimensions[ROW].height = 15
            COMBO_ROW1[combo_lbl] = ROW
            ROW += 1

        ROW += 2
        _title(ws, ROW, 1, col_30, '【구간 ②】 월별 누적 강수량 (mm) — 1월부터 해당 월까지 합계', h=24)
        ROW += 1
        write_hdr(ws, ROW, years)
        ROW += 1

        CUM_ROW = {}
        for m in range(1, 13):
            bg = C['light_blue'] if m % 2 == 0 else C['white']
            _dc(ws, ROW, 1, m, bg=bg, bold=True, nf='0')
            for yr in years:
                # 1~m월 ROUND된 값들의 합
                m_refs = '+'.join(f'{cl(yr)}{M_ROW[mm]}' for mm in range(1, m+1))
                _dc(ws, ROW, yr2col[yr], f'=ROUND(SUM({m_refs}),1)',
                    bg=bg, nf='#,##0.0')
            _dc(ws, ROW, col_10, yr_avg_formula(yr10, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            _dc(ws, ROW, col_20, yr_avg_formula(yr20, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            _dc(ws, ROW, col_30, yr_avg_formula(yr30, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            ws.row_dimensions[ROW].height = 15
            CUM_ROW[m] = ROW
            ROW += 1

        ROW += 1
        COMBO_ROW2 = {}
        for combo_lbl, months in SUMMER:
            _hc(ws, ROW, 1, combo_lbl, bg=C['mid_blue'], fg=C['white'], sz=9)
            for yr in years:
                m_refs = '+'.join(f'{cl(yr)}{M_ROW[m]}' for m in months)
                _dc(ws, ROW, yr2col[yr], f'=ROUND(SUM({m_refs}),1)',
                    bg=C['light_blue'], nf='#,##0.0')
            _dc(ws, ROW, col_10, yr_avg_formula(yr10, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            _dc(ws, ROW, col_20, yr_avg_formula(yr20, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            _dc(ws, ROW, col_30, yr_avg_formula(yr30, ROW), bg=C['light_orange'], bold=True, nf='#,##0.0')
            ws.row_dimensions[ROW].height = 15
            COMBO_ROW2[combo_lbl] = ROW
            ROW += 1

        ROW += 2
        _title(ws, ROW, 1, col_30, '【구간 ③】 연간 대비 누적 강수량 비율', h=24)
        ROW += 1
        write_hdr(ws, ROW, years)
        ROW += 1

        for m in range(1, 13):
            bg = C['light_blue'] if m % 2 == 0 else C['white']
            _dc(ws, ROW, 1, m, bg=bg, bold=True, nf='0')
            for yr in years:
                ann_ref = f'{cl(yr)}{ROW_ANN}'
                cum_ref = f'{cl(yr)}{CUM_ROW[m]}'
                _dc(ws, ROW, yr2col[yr],
                    f'=IFERROR(ROUND({cum_ref}/{ann_ref},3),"")',
                    bg=bg, nf='0.000')
            _dc(ws, ROW, col_10, yr_avg_formula(yr10, ROW), bg=C['light_orange'], bold=True, nf='0.000')
            _dc(ws, ROW, col_20, yr_avg_formula(yr20, ROW), bg=C['light_orange'], bold=True, nf='0.000')
            _dc(ws, ROW, col_30, yr_avg_formula(yr30, ROW), bg=C['light_orange'], bold=True, nf='0.000')
            ws.row_dimensions[ROW].height = 15
            ROW += 1

        ROW += 1
        for combo_lbl, months in SUMMER:
            _hc(ws, ROW, 1, combo_lbl, bg=C['mid_blue'], fg=C['white'], sz=9)
            for yr in years:
                ann_ref   = f'{cl(yr)}{ROW_ANN}'
                combo_ref = f'{cl(yr)}{COMBO_ROW1[combo_lbl]}'
                _dc(ws, ROW, yr2col[yr],
                    f'=IFERROR(ROUND({combo_ref}/{ann_ref},3),"")',
                    bg=C['light_blue'], nf='0.000')
            _dc(ws, ROW, col_10, yr_avg_formula(yr10, ROW), bg=C['light_orange'], bold=True, nf='0.000')
            _dc(ws, ROW, col_20, yr_avg_formula(yr20, ROW), bg=C['light_orange'], bold=True, nf='0.000')
            _dc(ws, ROW, col_30, yr_avg_formula(yr30, ROW), bg=C['light_orange'], bold=True, nf='0.000')
            ws.row_dimensions[ROW].height = 15
            ROW += 1

        # ── 누적강수량 차트 추가 ──
        ROW += 3
        self._add_cumulative_charts(ws, n, yr2col, col_10, col_20, col_30,
                                    M_ROW, CUM_ROW, ROW)




    def _fix_table_rels(self, filepath: str):
        """
        세 가지 문제를 수정합니다:
        1. 중복 autoFilter 태그 제거 (Table이 자체 포함)
        2. _rels Target 절대경로 → 상대경로 수정
        3. MAX(IF(...)) / MIN(IF(...)) 수식을 배열 수식(t="array")으로 마킹
           → Excel 365가 MAXIFS/MINIFS를 @MAXIFS로 자동 변환하는 문제 방지
        """
        import re as _re

        with zipfile.ZipFile(filepath, 'r') as zin:
            files = {n: zin.read(n) for n in zin.namelist()}

        changed = False

        # ── (1) 시트 XML에서 standalone <autoFilter> 제거 ──
        # Table이 있는 시트에서 <autoFilter .../> 독립 태그만 제거
        # (Table 자체의 autoFilter는 table1.xml 안에 있으므로 건드리지 않음)
        for name, data in list(files.items()):
            if not (name.startswith('xl/worksheets/sheet') and name.endswith('.xml')):
                continue
            xml = data.decode('utf-8')
            # tableParts가 있는 시트에서만 처리
            if '<tableParts' not in xml:
                continue
            # 시트 레벨 <autoFilter .../> 제거
            new_xml = _re.sub(r'<autoFilter[^>]*/>', '', xml)
            if new_xml != xml:
                files[name] = new_xml.encode('utf-8')
                changed = True

        # ── (2) _rels 파일의 절대경로 → 상대경로 수정 ──
        for name, data in list(files.items()):
            if not name.endswith('.rels'):
                continue
            xml = data.decode('utf-8')
            # /xl/tables/ → ../tables/  (worksheets/_rels/ 기준 상대경로)
            new_xml = xml.replace(
                'Target="/xl/tables/',
                'Target="../tables/'
            )
            # /xl/pivotTables/ → ../pivotTables/
            new_xml = new_xml.replace(
                'Target="/xl/pivotTables/',
                'Target="../pivotTables/'
            )
            if new_xml != xml:
                files[name] = new_xml.encode('utf-8')
                changed = True

        # ── (3) MAX(IF) / MIN(IF) → 배열 수식(t="array") 마킹 ──
        # Excel 365가 MAXIFS/MINIFS를 @MAXIFS로 자동변환하는 문제 방지
        for name, data in list(files.items()):
            if not (name.startswith('xl/worksheets/sheet') and name.endswith('.xml')):
                continue
            xml = data.decode('utf-8')
            if 'MAX(IF(' not in xml and 'MIN(IF(' not in xml:
                continue

            def mark_array_cell(m_cell):
                cell_str = m_cell.group(0)
                addr_m = _re.search(r'<c\b[^>]*?\br="([^"]+)"', cell_str)
                if not addr_m:
                    return cell_str
                addr = addr_m.group(1)
                def add_arr_attr(fm):
                    if 't="array"' in fm.group(0):
                        return fm.group(0)
                    return f'<f t="array" ref="{addr}">'
                return _re.sub(r'<f>', add_arr_attr, cell_str, count=1)

            new_xml = _re.sub(
                (r'<c\b[^>]*>(?:[^<]|<(?!/?c\b))*?'
                 r'<f>(?:[^<]*(?:MAX\(IF|MIN\(IF)[^<]*)</f>'
                 r'(?:[^<]|<(?!/?c\b))*?</c>'),
                mark_array_cell, xml
            )
            if new_xml != xml:
                files[name] = new_xml.encode('utf-8')
                changed = True

        if changed:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                for name, data in files.items():
                    zout.writestr(name, data)
            with open(filepath, 'wb') as f:
                f.write(buf.getvalue())

    # ─────────────────────────────────────────────────────────────
    # 차트 plot_area layout XML 직접 삽입
    # openpyxl이 chart layout을 write하지 않으므로 ZIP 수정 방식 사용
    # ─────────────────────────────────────────────────────────────
    def _inject_chart_layouts(self, filepath: str):
        """
        저장된 XLSX의 모든 차트 XML에 plotArea > layout 태그를 삽입합니다.
        차트 영역과 그림 영역 사이 약 10mm 여백을 확보합니다.
        """
        LAYOUT_XML = (
            '<layout><manualLayout>'
            '<layoutTarget val="inner"/>'
            '<xMode val="edge"/><yMode val="edge"/>'
            '<wMode val="edge"/><hMode val="edge"/>'
            '<x val="0.08"/><y val="0.12"/>'
            '<w val="0.84"/><h val="0.76"/>'
            '</manualLayout></layout>'
        )
        with zipfile.ZipFile(filepath, 'r') as zin:
            files = {n: zin.read(n) for n in zin.namelist()}

        for name, data in list(files.items()):
            if not (name.startswith('xl/charts/chart') and name.endswith('.xml')):
                continue
            xml = data.decode('utf-8')
            if '<layout>' in xml:
                continue
            new_xml = xml.replace('<plotArea>', f'<plotArea>{LAYOUT_XML}', 1)
            if new_xml != xml:
                files[name] = new_xml.encode('utf-8')

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in files.items():
                zout.writestr(name, data)
        with open(filepath, 'wb') as f:
            f.write(buf.getvalue())

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # PivotTable XML 삽입  (요구사항 6)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _inject_pivot_table(self, filepath: str, df: pd.DataFrame,
                             selected_cols: list):
        """
        저장된 XLSX에 실제 Excel PivotTable XML을 삽입합니다.
        소스: '원본 데이터' 시트의 기상데이터 Table
        위치: '피벗작업' 시트 A3
        """
        import xml.etree.ElementTree as ET

        with zipfile.ZipFile(filepath, 'r') as zin:
            files = {n: zin.read(n) for n in zin.namelist()}

        # ── 시트명 → 파일명 찾기 ──
        WB_NS  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
        PKG_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
        R_NS   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

        def find_sheet_file(sheet_name):
            wb_root = ET.fromstring(files['xl/workbook.xml'])
            rid = None
            for sh in wb_root.findall(f'.//{{{WB_NS}}}sheet'):
                if sh.get('name') == sheet_name:
                    rid = sh.get(f'{{{R_NS}}}id')
                    break
            if not rid:
                return None
            rels_root = ET.fromstring(files['xl/_rels/workbook.xml.rels'])
            for rel in rels_root.findall(f'{{{PKG_NS}}}Relationship'):
                if rel.get('Id') == rid:
                    return rel.get('Target')  # e.g. worksheets/sheet5.xml
            return None

        pivot_target  = find_sheet_file('피벗작업')
        if not pivot_target:
            return

        # 절대경로(/xl/worksheets/...) → 상대경로(worksheets/...) 정규화
        if pivot_target.startswith('/'):
            pivot_target = pivot_target.lstrip('/')
        # 'xl/worksheets/sheetN.xml' 또는 'worksheets/sheetN.xml' 통일
        if pivot_target.startswith('xl/'):
            pivot_target = pivot_target[3:]   # 'worksheets/sheetN.xml'


        # ── 필드 정의 ──
        headers = getattr(self, '_raw_col_headers', ['날짜','연도','월'])
        fields = []
        for h in headers:
            if h in ('날짜',):
                fields.append((h, 'date'))
            elif h in ('연도','월') or h in list(ELEMENT_LABELS.values()):
                fields.append((h, 'number'))
            else:
                fields.append((h, 'string'))
        n_fields = len(fields)

        # 행: 연도(인덱스1), 열: 월(인덱스2), 값: 강수량 or 기온
        row_idx  = 1  # 연도
        col_idx  = 2  # 월
        val_lbl  = ELEMENT_LABELS.get('precipitation', '강수량(mm)')
        val_idx  = next((i for i,(h,t) in enumerate(fields)
                         if h == val_lbl), n_fields - 1)

        # ── cacheId 결정 ──
        wb_root  = ET.fromstring(files['xl/workbook.xml'])
        existing = wb_root.findall(f'.//{{{WB_NS}}}pivotCache')
        cache_id = len(existing) + 1
        cache_rid = f'rIdPivotCache{cache_id}'

        # ── XML 생성 ──
        cache_fields_xml = '\n'.join(
            f'<cacheField name="{h}" numFmtId="{"14" if t=="date" else "0"}">'
            f'<sharedItems '
            f'{"containsDate=\"1\" containsNonDate=\"0\" " if t=="date" else ""}'
            f'containsString="{"0" if t!="string" else "1"}" '
            f'{"containsNumber=\"1\" " if t=="number" else ""}'
            f'count="0"/></cacheField>'
            for h, t in fields
        )

        cache_def_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
  r:id="rId1" refreshedBy="Python" refreshOnLoad="1" createdVersion="5"
  refreshedVersion="5" minRefreshableVersion="5" recordCount="0">
  <cacheSource type="worksheet">
    <worksheetSource ref="{self._raw_table_range}" sheet="📊 원본 데이터"/>
  </cacheSource>
  <cacheFields count="{n_fields}">
{cache_fields_xml}
  </cacheFields>
</pivotCacheDefinition>'''

        cache_rec_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0"/>'''

        cache_rels_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords"
    Target="pivotCacheRecords1.xml"/>
</Relationships>'''

        # pivotField 정의
        pf_list = []
        for i in range(n_fields):
            if i == row_idx:
                pf = ('<pivotField axis="axisRow" compact="0" outline="0" '
                      'subtotalTop="0" showAll="0" includeNewItemsInFilter="1">'
                      '<items count="1"><item t="default"/></items></pivotField>')
            elif i == col_idx:
                pf = ('<pivotField axis="axisCol" compact="0" outline="0" '
                      'subtotalTop="0" showAll="0" includeNewItemsInFilter="1">'
                      '<items count="1"><item t="default"/></items></pivotField>')
            elif i == val_idx:
                pf = ('<pivotField dataField="1" compact="0" outline="0" '
                      'subtotalTop="0" showAll="0" includeNewItemsInFilter="1">'
                      '<items count="1"><item t="default"/></items></pivotField>')
            else:
                pf = ('<pivotField compact="0" outline="0" subtotalTop="0" '
                      'showAll="0" includeNewItemsInFilter="1"/>')
            pf_list.append(pf)

        pivot_tbl_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  name="피벗 테이블1" cacheId="{cache_id}"
  applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0"
  applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1"
  dataCaption="값" updatedVersion="5" minRefreshableVersion="5"
  showDrill="1" useAutoFormatting="1" itemPrintTitles="1"
  createdVersion="5" indent="2" compact="0" outline="0"
  outlineData="0" multipleFieldFilters="0">
  <location ref="A3" firstHeaderRow="1" firstDataRow="2" firstDataCol="0"
    rowPageCount="1" colPageCount="1"/>
  <pivotFields count="{n_fields}">
{"".join(pf_list)}
  </pivotFields>
  <rowFields count="1"><field x="{row_idx}"/></rowFields>
  <colFields count="1"><field x="{col_idx}"/></colFields>
  <dataFields count="1">
    <dataField name="합계 : {val_lbl}" fld="{val_idx}"
      subtotal="sum" showDataAs="normal" baseField="0" baseItem="0"/>
  </dataFields>
</pivotTableDefinition>'''

        pivot_rels_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition"
    Target="../pivotCache/pivotCacheDefinition1.xml"/>
</Relationships>'''

        # ── [Content_Types].xml 수정 ──
        ct_xml  = files['[Content_Types].xml'].decode('utf-8')
        new_ct  = (
            '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.pivotCacheDefinition+xml"/>\n'
            '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.pivotCacheRecords+xml"/>\n'
            '<Override PartName="/xl/pivotTables/pivotTable1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.pivotTable+xml"/>'
        )
        ct_xml = ct_xml.replace('</Types>', new_ct + '\n</Types>')
        files['[Content_Types].xml'] = ct_xml.encode('utf-8')

        # ── xl/workbook.xml に pivotCaches 追加 ──
        wb_xml = files['xl/workbook.xml'].decode('utf-8')
        # workbook.xml에 r 네임스페이스가 선언되어 있는지 확인 후 처리
        pivot_caches_elem = (
            f'<pivotCaches><pivotCache cacheId="{cache_id}" '
            f'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            f'r:id="{cache_rid}"/></pivotCaches>')
        if '<pivotCaches>' not in wb_xml:
            wb_xml = wb_xml.replace('</workbook>', pivot_caches_elem + '</workbook>')
        files['xl/workbook.xml'] = wb_xml.encode('utf-8')

        # ── xl/_rels/workbook.xml.rels にキャッシュ relationship 追加 ──
        wb_rels = files['xl/_rels/workbook.xml.rels'].decode('utf-8')
        new_rel = (
            f'<Relationship Id="{cache_rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument'
            f'/2006/relationships/pivotCacheDefinition" '
            f'Target="pivotCache/pivotCacheDefinition1.xml"/>')
        wb_rels = wb_rels.replace('</Relationships>', new_rel + '</Relationships>')
        files['xl/_rels/workbook.xml.rels'] = wb_rels.encode('utf-8')

        # ── 피벗 시트 .rels 파일 생성/수정 ──
        # pivot_target = 'worksheets/sheetN.xml'
        _sheet_fname = pivot_target.replace('worksheets/', '')  # 'sheetN.xml'
        sheet_rels_key = f'xl/worksheets/_rels/{_sheet_fname}.rels'
        if sheet_rels_key in files:
            sr = files[sheet_rels_key].decode('utf-8')
            sr = sr.replace('</Relationships>',
                            '<Relationship Id="rIdPivotTable1" '
                            'Type="http://schemas.openxmlformats.org/officeDocument'
                            '/2006/relationships/pivotTable" '
                            'Target="../pivotTables/pivotTable1.xml"/>'
                            '</Relationships>')
        else:
            sr = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                  '<Relationships xmlns="http://schemas.openxmlformats.org'
                  '/package/2006/relationships">\n'
                  '<Relationship Id="rIdPivotTable1" '
                  'Type="http://schemas.openxmlformats.org/officeDocument'
                  '/2006/relationships/pivotTable" '
                  'Target="../pivotTables/pivotTable1.xml"/>'
                  '\n</Relationships>')
        files[sheet_rels_key] = sr.encode('utf-8')

        # ── 새 XML 파일 추가 ──
        files['xl/pivotCache/pivotCacheDefinition1.xml'] = cache_def_xml.encode('utf-8')
        files['xl/pivotCache/pivotCacheRecords1.xml']    = cache_rec_xml.encode('utf-8')
        files['xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels'] = cache_rels_xml.encode('utf-8')
        files['xl/pivotTables/pivotTable1.xml']          = pivot_tbl_xml.encode('utf-8')
        files['xl/pivotTables/_rels/pivotTable1.xml.rels'] = pivot_rels_xml.encode('utf-8')

        # ── 파일 재작성 ──
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, content in files.items():
                zout.writestr(name, content)
        with open(filepath, 'wb') as f:
            f.write(buf.getvalue())

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 원본 데이터2  — 세로형(Long Format) 데이터
    #   열 구성: 날짜 | 연도 | 월 | 관측소명 | 항목 | Data
    #   원본 데이터의 모든 기상요소를 '항목' 열 + 'Data' 1열로 변환 (melt)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_raw2(self, wb, df: pd.DataFrame, selected_cols: list):
        ws = wb.create_sheet("원본 데이터2")

        avail = [c for c in selected_cols if c in df.columns and c in ELEMENT_LABELS]
        if not avail:
            return

        # ── pandas melt로 세로 변환 ──
        id_cols = ['date']
        if 'station_name' in df.columns:
            id_cols.append('station_name')

        sub = df[id_cols + avail].copy()
        sub['_year']  = sub['date'].dt.year
        sub['_month'] = sub['date'].dt.month
        sub['_date']  = sub['date'].dt.strftime('%Y-%m-%d')

        # melt: 기상요소 컬럼들을 '항목' + 'Data' 두 열로 변환
        value_cols = avail
        id_vars    = ['_date', '_year', '_month']
        if 'station_name' in sub.columns:
            id_vars.append('station_name')

        melted = sub.melt(
            id_vars=id_vars,
            value_vars=value_cols,
            var_name='_item_key',
            value_name='Data'
        )

        # 항목명을 한글로 변환
        melted['항목'] = melted['_item_key'].map(ELEMENT_LABELS)

        # 열 순서 정리
        out_cols = ['_date', '_year', '_month']
        if 'station_name' in melted.columns:
            out_cols.append('station_name')
        out_cols += ['항목', 'Data']
        melted = melted[out_cols].copy()

        # 날짜·항목 기준 정렬
        melted = melted.sort_values(['_date', '_item_key'] if '_item_key' in melted.columns
                                    else ['_date', '항목']).reset_index(drop=True)

        n_rows  = len(melted)
        n_cols  = len(out_cols)

        # ── 헤더 ──
        headers = ['날짜', '연도', '월']
        if 'station_name' in melted.columns:
            headers.append('관측소명')
        headers += ['항목', 'Data']

        for j, h in enumerate(headers, 1):
            _hc(ws, 1, j, h, bg=C['dark_blue'], sz=10)
            ws.column_dimensions[get_column_letter(j)].width = 16
        ws.row_dimensions[1].height = 20

        # ── 데이터 행 ──
        for i, (_, row_d) in enumerate(melted.iterrows(), start=2):
            bg = C['light_blue'] if i % 2 == 0 else C['white']
            col = 1

            # 날짜 (문자열)
            _dc(ws, i, col, row_d['_date'], bg=bg, sz=9); col += 1
            # 연도 (수식)
            _dc(ws, i, col, f'=YEAR(A{i})', bg=bg, sz=9, nf='0'); col += 1
            # 월 (수식)
            _dc(ws, i, col, f'=MONTH(A{i})', bg=bg, sz=9, nf='0'); col += 1

            if 'station_name' in melted.columns:
                _dc(ws, i, col, row_d.get('station_name', ''), bg=bg, sz=9); col += 1

            # 항목
            _dc(ws, i, col, row_d['항목'], bg=bg, sz=9, align='left'); col += 1

            # Data
            val = row_d['Data']
            if isinstance(val, float) and np.isnan(val):
                val = None
            _dc(ws, i, col, val, bg=bg, sz=9,
                nf='#,##0.0' if isinstance(val, float) else None)

            ws.row_dimensions[i].height = 14

        # ── Excel Table 생성 ──
        last_col  = get_column_letter(len(headers))
        tbl_range = f'A1:{last_col}{n_rows + 1}'
        # 피벗테이블2 소스 정보 저장
        self._raw2_headers = headers
        self._raw2_nrows   = n_rows
        tbl2 = Table(displayName="기상데이터_세로형", ref=tbl_range)
        tbl2.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True,   showColumnStripes=False
        )
        ws.add_table(tbl2)
        ws.freeze_panes = 'A2'

        # 시트 설명 추가
        _note(ws, n_rows + 3, 1, len(headers),
              f'※ 총 {n_rows:,}행 ({len(avail)}개 기상요소 × {n_rows//len(avail):,}일) '
              '| [삽입→피벗테이블]로 다양한 분석 가능')

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 누적강수량 차트 (_sheet_cumulative_precip 호출용)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _add_cumulative_charts(self, ws, n_year, yr2col, col_10, col_20, col_30,
                                M_ROW, CUM_ROW, chart_start_row):
        """
        누적강수량 분석 3종 차트:
          ①  월별 강수량 평균 막대 (구간① 데이터 활용)
          ②  연도별 누적강수량 꺾은선  (구간② 데이터, 각 연도 + 평균)
          ③  연간 대비 누적비율 꺾은선 (구간③ 데이터, 각 연도 + 평균)
        """
        # ── 고정 행 위치 계산 ──
        # 구간① : 헤더=2, 월데이터=3~14, 합계=15
        # 구간② : 헤더=26, 누적데이터=27~38
        # 구간③ : 헤더=48, 비율데이터=49~60
        HDR1 = 2;  D1_S = 3;   D1_E = 14
        HDR2 = 25; D2_S = 26;  D2_E = 37
        HDR3 = 47; D3_S = 48;  D3_E = 59

        # 카테고리 (월 라벨) — 구간①의 A열 참조
        cats1 = Reference(ws, min_col=1, min_row=D1_S, max_row=D1_E)
        cats2 = Reference(ws, min_col=1, min_row=D2_S, max_row=D2_E)
        cats3 = Reference(ws, min_col=1, min_row=D3_S, max_row=D3_E)

        W = 22; H = 13
        ROW = chart_start_row

        # ── ① 월별 강수량 평균 막대 차트 ──
        ch1 = BarChart()
        ch1.type = 'col'
        ch1.title = "월별 강수량 (평균)"
        ch1.style = 10
        ch1.y_axis.title = "강수량 (mm)"
        ch1.x_axis.title = "월"
        ch1.width, ch1.height = W, H

        for avg_col, color in [(col_10, '4472C4'), (col_20, 'ED7D31'), (col_30, '70AD47')]:
            d = Reference(ws, min_col=avg_col, min_row=HDR1, max_row=D1_E)
            ch1.add_data(d, titles_from_data=True)
            ch1.series[-1].graphicalProperties.solidFill = color
        ch1.set_categories(cats1)
        _set_chart_layout(ch1)
        ws.add_chart(ch1, f'A{ROW}')

        # ── ② 누적강수량 꺾은선 차트 ──
        ch2 = LineChart()
        ch2.title = "월별 누적 강수량 (연도별)"
        ch2.style = 10
        ch2.y_axis.title = "누적강수량 (mm)"
        ch2.x_axis.title = "월"
        ch2.width, ch2.height = W, H

        # 모든 연도 (연회색 thin)
        for yr, col in yr2col.items():
            d = Reference(ws, min_col=col, min_row=HDR2, max_row=D2_E)
            ch2.add_data(d, titles_from_data=True)
            s = ch2.series[-1]
            s.graphicalProperties.line.solidFill = 'BBBBBB'
            s.graphicalProperties.line.width = 9525   # 1pt

        # 평균선 (굵고 선명하게)
        for avg_col, color, width in [
            (col_10, 'FF0000', 28000),
            (col_20, 'ED7D31', 22000),
            (col_30, '0070C0', 28000),
        ]:
            d = Reference(ws, min_col=avg_col, min_row=HDR2, max_row=D2_E)
            ch2.add_data(d, titles_from_data=True)
            s = ch2.series[-1]
            s.graphicalProperties.line.solidFill = color
            s.graphicalProperties.line.width = width

        ch2.set_categories(cats2)
        _set_chart_layout(ch2)
        ws.add_chart(ch2, f'L{ROW}')
        ROW += 25

        # ── ③ 누적비율 꺾은선 차트 ──
        ch3 = LineChart()
        ch3.title = "연간 대비 누적 강수량 비율"
        ch3.style = 10
        ch3.y_axis.title = "누적비율"
        ch3.x_axis.title = "월"
        ch3.width, ch3.height = W * 2 + 2, H

        for yr, col in yr2col.items():
            d = Reference(ws, min_col=col, min_row=HDR3, max_row=D3_E)
            ch3.add_data(d, titles_from_data=True)
            s = ch3.series[-1]
            s.graphicalProperties.line.solidFill = 'BBBBBB'
            s.graphicalProperties.line.width = 9525

        for avg_col, color, width in [
            (col_10, 'FF0000', 28000),
            (col_20, 'ED7D31', 22000),
            (col_30, '0070C0', 28000),
        ]:
            d = Reference(ws, min_col=avg_col, min_row=HDR3, max_row=D3_E)
            ch3.add_data(d, titles_from_data=True)
            s = ch3.series[-1]
            s.graphicalProperties.line.solidFill = color
            s.graphicalProperties.line.width = width

        ch3.set_categories(cats3)
        _set_chart_layout(ch3)
        ws.add_chart(ch3, f'A{ROW}')

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 강우일수 분석 시트 (NEW)
    # 카테고리: 무강우 / <3 / 3~6 / 6~10 / 10~20 / 20~50 / >=50 / 강우일수
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_rainfall_days(self, wb, df, selected_cols):
        if 'precipitation' not in df.columns or 'precipitation' not in selected_cols:
            return

        ws = wb.create_sheet("☔ 강우일수 분석")
        ws.sheet_view.showGridLines = True
        ws.freeze_panes = 'B3'

        yc    = self._col.get('year',          'B')
        pc    = self._col.get('precipitation')
        years = self._years
        RAW   = RAW_SHEET

        # 보조 컬럼 위치: K=총일수, L=강우일수(>0), M=>=3, N=>=6, O=>=10, P=>=20, Q=>=50
        AUX_START = 11  # K열

        # ── 타이틀 ──
        ROW = 1
        _title(ws, ROW, 1, 10, '강우일수 분석 — 강수량 규모별 발생일수', h=26)
        ROW += 1

        # ── 헤더 2행 구조 ──
        hdr_items = [
            ('연도',   10),
            ('무강우일', 9),
            ('<3mm',   9),
            ('3~6mm',  9),
            ('6~10mm', 9),
            ('10~20mm', 9),
            ('20~50mm', 9),
            ('≥50mm',  9),
            ('강우일수', 10),
        ]
        for j, (lbl, w) in enumerate(hdr_items, 1):
            _hc(ws, ROW, j, lbl, bg=C['dark_blue'], sz=9)
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.row_dimensions[ROW].height = 18
        ROW += 1

        data_start = ROW

        for i, yr in enumerate(years):
            bg  = C['light_blue'] if i % 2 == 0 else C['white']
            aux = AUX_START   # K열부터 보조 데이터
            kc  = get_column_letter(aux)    # 총일수
            lc  = get_column_letter(aux+1)  # 강우일수 >0
            mc  = get_column_letter(aux+2)  # >=3
            nc  = get_column_letter(aux+3)  # >=6
            oc  = get_column_letter(aux+4)  # >=10
            qc  = get_column_letter(aux+5)  # >=20
            rc  = get_column_letter(aux+6)  # >=50

            # 보조 컬럼에 COUNTIFS 수식 작성 (각 연도마다)
            ws.cell(row=ROW, column=aux).value   = f'=COUNTIFS({RAW}!{yc}:{yc},{yr})'
            ws.cell(row=ROW, column=aux+1).value = f'=COUNTIFS({RAW}!{yc}:{yc},{yr},{RAW}!{pc}:{pc},">0")'
            ws.cell(row=ROW, column=aux+2).value = f'=COUNTIFS({RAW}!{yc}:{yc},{yr},{RAW}!{pc}:{pc},">=3")'
            ws.cell(row=ROW, column=aux+3).value = f'=COUNTIFS({RAW}!{yc}:{yc},{yr},{RAW}!{pc}:{pc},">=6")'
            ws.cell(row=ROW, column=aux+4).value = f'=COUNTIFS({RAW}!{yc}:{yc},{yr},{RAW}!{pc}:{pc},">=10")'
            ws.cell(row=ROW, column=aux+5).value = f'=COUNTIFS({RAW}!{yc}:{yc},{yr},{RAW}!{pc}:{pc},">=20")'
            ws.cell(row=ROW, column=aux+6).value = f'=COUNTIFS({RAW}!{yc}:{yc},{yr},{RAW}!{pc}:{pc},">=50")'

            # 보조 열 서식 (회색 작게)
            for ci in range(aux, aux+7):
                ws.column_dimensions[get_column_letter(ci)].width = 8
                ws.cell(row=ROW, column=ci).font = Font(name=FONT, size=8, color='888888')

            # 표시 컬럼 (수식 참조)
            _dc(ws, ROW, 1, yr, bg=bg, bold=True, nf='0')
            # 무강우 = 총일수 - 강우일수
            _dc(ws, ROW, 2, f'={kc}{ROW}-{lc}{ROW}', bg=bg, nf='0')
            # <3mm = 강우일수 - >=3mm
            _dc(ws, ROW, 3, f'={lc}{ROW}-{mc}{ROW}', bg=bg, nf='0')
            # 3~6mm = >=3mm - >=6mm
            _dc(ws, ROW, 4, f'={mc}{ROW}-{nc}{ROW}', bg=bg, nf='0')
            # 6~10mm = >=6mm - >=10mm
            _dc(ws, ROW, 5, f'={nc}{ROW}-{oc}{ROW}', bg=bg, nf='0')
            # 10~20mm = >=10mm - >=20mm
            _dc(ws, ROW, 6, f'={oc}{ROW}-{qc}{ROW}', bg=bg, nf='0')
            # 20~50mm = >=20mm - >=50mm
            _dc(ws, ROW, 7, f'={qc}{ROW}-{rc}{ROW}', bg=bg, nf='0')
            # >=50mm
            _dc(ws, ROW, 8, f'={rc}{ROW}', bg=bg, nf='0')
            # 강우일수
            _dc(ws, ROW, 9, f'={lc}{ROW}', bg=bg, bold=True, nf='0')
            ws.row_dimensions[ROW].height = 15
            ROW += 1

        data_end = ROW - 1

        # ── 집계행 (평균/최대/최소) ──
        for s_lbl, s_fn, s_bg, s_hbg, s_fg in [
            ('평균','AVERAGE',C['light_yellow'],C['yellow'],   C['dark_blue']),
            ('최대','MAX',    C['light_green'], C['mid_green'],C['white']),
            ('최소','MIN',    C['light_blue'],  C['mid_blue'], C['white']),
        ]:
            _hc(ws, ROW, 1, s_lbl, bg=s_hbg, fg=s_fg, sz=9)
            for j in range(2, 10):
                cl2 = get_column_letter(j)
                _dc(ws, ROW, j,
                    f'=ROUND({s_fn}({cl2}{data_start}:{cl2}{data_end}),1)',
                    bg=s_bg, bold=True, nf='0.0')
            ws.row_dimensions[ROW].height = 15
            ROW += 1

        ROW += 3

        # ── 차트: 강우일수 규모별 누적 막대 ──
        # 보조 데이터 (차트용 집계)
        AUX_C = AUX_START + 8   # 차트 보조 열 시작
        cat_labels = ['무강우','<3mm','3~6mm','6~10mm','10~20mm','20~50mm','≥50mm']
        ws.cell(row=1, column=AUX_C, value='구분')
        ws.cell(row=1, column=AUX_C+1, value='평균일수')
        for j2, lbl in enumerate(cat_labels):
            ws.cell(row=2+j2, column=AUX_C, value=lbl)
            col_l = get_column_letter(j2+2)
            ws.cell(row=2+j2, column=AUX_C+1,
                    value=f'=ROUND(AVERAGE({col_l}{data_start}:{col_l}{data_end}),1)')

        ch_rain = BarChart()
        ch_rain.type  = 'bar'
        ch_rain.title = "강수량 규모별 평균 발생일수"
        ch_rain.style = 10
        ch_rain.x_axis.title = "일수 (일)"
        ch_rain.width, ch_rain.height = 22, 13

        d_ref = Reference(ws, min_col=AUX_C+1, max_col=AUX_C+1, min_row=1, max_row=8)
        c_ref = Reference(ws, min_col=AUX_C,   min_row=2, max_row=8)
        ch_rain.add_data(d_ref, titles_from_data=True)
        ch_rain.set_categories(c_ref)
        ch_rain.series[0].graphicalProperties.solidFill = '4472C4'
        _set_chart_layout(ch_rain)
        ws.add_chart(ch_rain, f'A{ROW}')

        # ── 차트: 강우일수 연도별 추이 ──
        ch_trend = LineChart()
        ch_trend.title = "연도별 강우일수 추이"
        ch_trend.style = 10
        ch_trend.y_axis.title = "일수 (일)"
        ch_trend.x_axis.title = "연도"
        ch_trend.width, ch_trend.height = 22, 13

        # 연도 헤더
        AUX_T = AUX_C + 3
        ws.cell(row=1, column=AUX_T, value='연도')
        ws.cell(row=1, column=AUX_T+1, value='강우일수')
        ws.cell(row=1, column=AUX_T+2, value='무강우일')
        for i, yr in enumerate(years):
            ws.cell(row=2+i, column=AUX_T, value=yr)
            ws.cell(row=2+i, column=AUX_T+1, value=f'=I{data_start+i}')
            ws.cell(row=2+i, column=AUX_T+2, value=f'=B{data_start+i}')

        d_t1 = Reference(ws, min_col=AUX_T+1, max_col=AUX_T+2, min_row=1, max_row=1+len(years))
        c_t1 = Reference(ws, min_col=AUX_T,   min_row=2, max_row=1+len(years))
        ch_trend.add_data(d_t1, titles_from_data=True)
        ch_trend.set_categories(c_t1)
        ch_trend.series[0].graphicalProperties.line.solidFill = '2E75B6'
        ch_trend.series[1].graphicalProperties.line.solidFill = 'ED7D31'
        _set_chart_layout(ch_trend)
        ws.add_chart(ch_trend, f'L{ROW}')

        # 보조 열 숨기기 설정 (선택적으로 숨김 - 필요시 사용자가 숨길 수 있음)
        _note(ws, ROW-1, 1, 9, '※ K열 이후는 COUNTIFS 집계 보조 열입니다.')

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 월_기온&강수량 분석 시트 (NEW)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_monthly_tp(self, wb, df, selected_cols):
        ws = wb.create_sheet("📅 월_기온&강수량 분석")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'B4'

        yc    = self._col.get('year',  'B')
        mc    = self._col.get('month', 'C')
        pc    = self._col.get('precipitation')
        tc    = self._col.get('temp_avg')
        tmaxc = self._col.get('temp_max')
        tminc = self._col.get('temp_min')
        wsc   = self._col.get('wind_speed')
        wmaxc = self._col.get('wind_max')
        years = self._years
        RAW   = RAW_SHEET
        n_yr  = len(years)

        def avg_f(col, yr):
            return f'=ROUND(AVERAGEIFS({RAW}!{col}:{col},{RAW}!{yc}:{yc},{yr}),1)'

        def sum_f(col, yr):
            return f'=ROUND(SUMIFS({RAW}!{col}:{col},{RAW}!{yc}:{yc},{yr}),1)'

        def avg_mo_f(col, m):
            return f'=ROUND(AVERAGEIFS({RAW}!{col}:{col},{RAW}!{mc}:{mc},{m}),1)'

        def sum_mo_yr_f(col, yr, m):
            return (f'=ROUND(SUMIFS({RAW}!{col}:{col},'
                    f'{RAW}!{yc}:{yc},{yr},{RAW}!{mc}:{mc},{m}),1)')

        def avg_mo_yr_f(col, yr, m):
            return (f'=ROUND(AVERAGEIFS({RAW}!{col}:{col},'
                    f'{RAW}!{yc}:{yc},{yr},{RAW}!{mc}:{mc},{m}),1)')

        ROW = 1

        # ════════════════════════════════════════════════════
        # 표 1: 연도별 연간 기상 통계 (columns = years)
        # ════════════════════════════════════════════════════
        _title(ws, ROW, 1, n_yr+2, '연도별 연간 기상 통계', h=26)
        ROW += 1
        _note(ws, ROW, 1, n_yr+2, '※ 수식 기반 — 원본 데이터 변경 시 자동 갱신')
        ROW += 1

        # 헤더: 항목 | 연도들 | 평균
        _hc(ws, ROW, 1, '항목', bg=C['dark_blue'], sz=10)
        for j, yr in enumerate(years, 2):
            _hc(ws, ROW, j, str(yr), bg=C['dark_blue'], sz=9)
            ws.column_dimensions[get_column_letter(j)].width = 9
        _hc(ws, ROW, n_yr+2, '평균', bg=C['orange'], sz=10)
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions[get_column_letter(n_yr+2)].width = 10
        ws.row_dimensions[ROW].height = 18
        ROW += 1

        t1_rows = []
        stat_defs = []
        if tc:
            stat_defs.append(('연평균기온(℃)',    tc,    'avg', '0.0'))
            stat_defs.append(('평균최고기온(℃)',   tmaxc, 'avg', '0.0'))
            stat_defs.append(('평균최저기온(℃)',   tminc, 'avg', '0.0'))
        if pc:
            stat_defs.append(('월평균강수량(mm)',  pc,    'mon', '#,##0.0'))
            stat_defs.append(('연강수량(mm)',      pc,    'sum', '#,##0.0'))
        if wsc:
            stat_defs.append(('평균풍속(m/s)',     wsc,   'avg', '0.0'))
        if wmaxc:
            stat_defs.append(('최대풍속(m/s)',     wmaxc, 'avg', '0.0'))

        for i, (lbl, col_k, fn_t, nf) in enumerate(stat_defs):
            if not col_k: continue
            bg = C['light_blue'] if i % 2 == 0 else C['white']
            _dc(ws, ROW, 1, lbl, bg=bg, bold=True, align='left')
            vals_row = []
            for j, yr in enumerate(years, 2):
                if fn_t == 'avg':
                    f = avg_f(col_k, yr)
                elif fn_t == 'sum':
                    f = sum_f(col_k, yr)
                else:  # 'mon' = 월평균강수량
                    f = f'=ROUND(SUMIFS({RAW}!{col_k}:{col_k},{RAW}!{yc}:{yc},{yr})/12,1)'
                _dc(ws, ROW, j, f, bg=bg, nf=nf)
                vals_row.append(get_column_letter(j))
            # 평균열
            b_l = get_column_letter(2); e_l = get_column_letter(n_yr+1)
            _dc(ws, ROW, n_yr+2,
                f'=ROUND(AVERAGE({b_l}{ROW}:{e_l}{ROW}),1)',
                bg=C['light_orange'], bold=True, nf=nf)
            ws.row_dimensions[ROW].height = 16
            t1_rows.append(ROW)
            ROW += 1

        ROW += 3

        # ════════════════════════════════════════════════════
        # 표 2: 월별 강수량 (years × 12months)
        # ════════════════════════════════════════════════════
        if pc:
            _title(ws, ROW, 1, 15, '월별 강수량 분석 (mm)', h=22, bg=C['mid_blue'])
            ROW += 1

            # 헤더
            _hc(ws, ROW, 1, '연도', bg=C['dark_blue'], sz=10)
            for m in range(1, 13):
                _hc(ws, ROW, m+1, f'{m}월', bg=C['dark_blue'], sz=9)
                ws.column_dimensions[get_column_letter(m+1)].width = 9
            _hc(ws, ROW, 14, '합계', bg=C['orange'], sz=10)
            ws.column_dimensions[get_column_letter(14)].width = 10
            ws.row_dimensions[ROW].height = 18
            ROW += 1

            t2_data_s = ROW
            for i, yr in enumerate(years):
                bg = C['light_blue'] if i % 2 == 0 else C['white']
                _dc(ws, ROW, 1, yr, bg=bg, bold=True, nf='0')
                for m in range(1, 13):
                    _dc(ws, ROW, m+1, sum_mo_yr_f(pc, yr, m), bg=bg, nf='#,##0.0')
                # 합계 = sum of month cells
                b_l = get_column_letter(2); e_l = get_column_letter(13)
                _dc(ws, ROW, 14, f'=ROUND(SUM({b_l}{ROW}:{e_l}{ROW}),1)',
                    bg=bg, bold=True, nf='#,##0.0')
                ws.row_dimensions[ROW].height = 15
                ROW += 1

            t2_data_e = ROW - 1

            for s_lbl, s_fn, s_bg, s_hbg, s_fg in [
                ('평균','AVERAGE',C['light_yellow'],C['yellow'],   C['dark_blue']),
                ('최대','MAX',    C['light_green'], C['mid_green'],C['white']),
                ('최소','MIN',    C['light_blue'],  C['mid_blue'], C['white']),
            ]:
                _hc(ws, ROW, 1, s_lbl, bg=s_hbg, fg=s_fg, sz=9)
                for j in range(2, 15):
                    cl2 = get_column_letter(j)
                    _dc(ws, ROW, j,
                        f'=ROUND({s_fn}({cl2}{t2_data_s}:{cl2}{t2_data_e}),1)',
                        bg=s_bg, bold=True, nf='#,##0.0')
                ws.row_dimensions[ROW].height = 15
                ROW += 1

            # 평균행 위치 저장 (차트용)
            t2_avg_row = ROW - 3
            ROW += 3

        # ════════════════════════════════════════════════════
        # 표 3: 월별 기온 (years × 12months)
        # ════════════════════════════════════════════════════
        if tc:
            _title(ws, ROW, 1, 14, '월별 평균기온 분석 (℃)', h=22, bg=C['mid_blue'])
            ROW += 1

            _hc(ws, ROW, 1, '연도', bg=C['dark_blue'], sz=10)
            for m in range(1, 13):
                _hc(ws, ROW, m+1, f'{m}월', bg=C['dark_blue'], sz=9)
            _hc(ws, ROW, 14, '연평균', bg=C['orange'], sz=10)
            ws.row_dimensions[ROW].height = 18
            ROW += 1

            t3_data_s = ROW
            for i, yr in enumerate(years):
                bg = C['light_blue'] if i % 2 == 0 else C['white']
                _dc(ws, ROW, 1, yr, bg=bg, bold=True, nf='0')
                for m in range(1, 13):
                    _dc(ws, ROW, m+1, avg_mo_yr_f(tc, yr, m), bg=bg, nf='0.0')
                b_l = get_column_letter(2); e_l = get_column_letter(13)
                _dc(ws, ROW, 14, f'=ROUND(AVERAGE({b_l}{ROW}:{e_l}{ROW}),1)',
                    bg=bg, bold=True, nf='0.0')
                ws.row_dimensions[ROW].height = 15
                ROW += 1

            t3_data_e = ROW - 1

            for s_lbl, s_fn, s_bg, s_hbg, s_fg in [
                ('평균','AVERAGE',C['light_yellow'],C['yellow'],   C['dark_blue']),
                ('최대','MAX',    C['light_green'], C['mid_green'],C['white']),
                ('최소','MIN',    C['light_blue'],  C['mid_blue'], C['white']),
            ]:
                _hc(ws, ROW, 1, s_lbl, bg=s_hbg, fg=s_fg, sz=9)
                for j in range(2, 15):
                    cl2 = get_column_letter(j)
                    _dc(ws, ROW, j,
                        f'=ROUND({s_fn}({cl2}{t3_data_s}:{cl2}{t3_data_e}),1)',
                        bg=s_bg, bold=True, nf='0.0')
                ws.row_dimensions[ROW].height = 15
                ROW += 1

            t3_avg_row = ROW - 3
            ROW += 3

        # ════════════════════════════════════════════════════
        # 차트 A: 월별 강수량 (막대 + 평균선)
        # ════════════════════════════════════════════════════
        AUX = max(16, n_yr + 4)  # 차트 보조 데이터 (표1 끝 열 이후)

        if pc:
            # 보조: 월별 평균강수량 (전체 기간 AVERAGEIFS)
            ws.cell(row=1, column=AUX, value='월')
            ws.cell(row=1, column=AUX+1, value='월평균강수량(mm)')
            for m in range(1, 13):
                ws.cell(row=1+m, column=AUX, value=f'{m}월')
                ws.cell(row=1+m, column=AUX+1,
                        value=f'=ROUND(SUMIFS({RAW}!{pc}:{pc},{RAW}!{mc}:{mc},{m})'
                              f'/{n_yr},1)')

            ch_p = BarChart()
            ch_p.type = 'col'
            ch_p.title = "월별 강수량 (전체 기간 평균)"
            ch_p.style = 10
            ch_p.y_axis.title = "강수량 (mm)"
            ch_p.x_axis.title = "월"
            ch_p.width, ch_p.height = 24, 14

            d_p = Reference(ws, min_col=AUX+1, max_col=AUX+1, min_row=1, max_row=13)
            c_p = Reference(ws, min_col=AUX,   min_row=2, max_row=13)
            ch_p.add_data(d_p, titles_from_data=True)
            ch_p.set_categories(c_p)
            ch_p.series[0].graphicalProperties.solidFill = '4472C4'
            _set_chart_layout(ch_p)
            ws.add_chart(ch_p, f'A{ROW}')

        # ════════════════════════════════════════════════════
        # 차트 B: 월별 기온 (꺾은선, 각 연도 + 평균)
        # ════════════════════════════════════════════════════
        if tc:
            AUX2 = AUX + 3
            ws.cell(row=1, column=AUX2, value='월')
            ws.cell(row=1, column=AUX2+1, value='평균기온(℃)')
            ws.cell(row=1, column=AUX2+2, value='평균최고기온(℃)')
            ws.cell(row=1, column=AUX2+3, value='평균최저기온(℃)')
            for m in range(1, 13):
                ws.cell(row=1+m, column=AUX2, value=f'{m}월')
                ws.cell(row=1+m, column=AUX2+1, value=avg_mo_f(tc, m))
                if tmaxc:
                    ws.cell(row=1+m, column=AUX2+2, value=avg_mo_f(tmaxc, m))
                if tminc:
                    ws.cell(row=1+m, column=AUX2+3, value=avg_mo_f(tminc, m))

            ch_t = LineChart()
            ch_t.title = "월별 기온 (전체 기간 평균)"
            ch_t.style = 10
            ch_t.y_axis.title = "기온 (℃)"
            ch_t.x_axis.title = "월"
            ch_t.width, ch_t.height = 24, 14

            n_t_cols = 1 + (1 if tmaxc else 0) + (1 if tminc else 0)
            d_t = Reference(ws, min_col=AUX2+1, max_col=AUX2+n_t_cols, min_row=1, max_row=13)
            c_t = Reference(ws, min_col=AUX2,   min_row=2, max_row=13)
            ch_t.add_data(d_t, titles_from_data=True)
            ch_t.set_categories(c_t)
            for k, color in enumerate(['4472C4','ED7D31','70AD47']):
                if k < len(ch_t.series):
                    ch_t.series[k].graphicalProperties.line.solidFill = color
                    ch_t.series[k].graphicalProperties.line.width = 22000
            _set_chart_layout(ch_t)
            ws.add_chart(ch_t, f'M{ROW}')

        # 보조 열 숨기기 없음 (요구사항 1 반영)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 원본 데이터2 피벗테이블 삽입 (NEW)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _inject_pivot_table2(self, filepath: str):
        """
        '원본 데이터2'(세로형) 시트를 소스로 하는 PivotTable을 '피벗작업2' 시트에 삽입.
        날짜·연도·월·관측소명·항목·Data 6개 필드가 필드 목록에 표시됨.
        """
        import xml.etree.ElementTree as ET

        with zipfile.ZipFile(filepath, 'r') as zin:
            files = {n: zin.read(n) for n in zin.namelist()}

        WB_NS  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
        PKG_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
        R_NS   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

        def find_sheet_file(sheet_name):
            wb_root = ET.fromstring(files['xl/workbook.xml'])
            rid = None
            for sh in wb_root.findall(f'.//{{{WB_NS}}}sheet'):
                if sh.get('name') == sheet_name:
                    rid = sh.get(f'{{{R_NS}}}id')
                    break
            if not rid:
                return None
            rels_root = ET.fromstring(files['xl/_rels/workbook.xml.rels'])
            for rel in rels_root.findall(f'{{{PKG_NS}}}Relationship'):
                if rel.get('Id') == rid:
                    target = rel.get('Target')
                    if target.startswith('/'):
                        target = target.lstrip('/')
                    if target.startswith('xl/'):
                        target = target[3:]
                    return target
            return None

        pivot2_target = find_sheet_file('피벗작업2')
        if not pivot2_target:
            return

        # 원본 데이터2 Table 범위 확인
        headers2 = getattr(self, '_raw2_headers', ['날짜','연도','월','관측소명','항목','Data'])
        n_rows2   = getattr(self, '_raw2_nrows',  1)
        n_cols2   = len(headers2)
        last_col2 = get_column_letter(n_cols2)
        table2_ref = f'A1:{last_col2}{n_rows2 + 1}'

        # cacheId = 2
        CACHE_ID = 2
        CACHE_RID = 'rIdPivotCache2'

        # 필드 XML 생성
        cache_fields_xml = '\n'.join(
            f'<cacheField name="{h}" numFmtId="0">'
            f'<sharedItems containsString="1" count="0"/></cacheField>'
            for h in headers2
        )

        # 값 필드: 'Data' (마지막)
        val_idx = n_cols2 - 1
        pf_items = []
        for i in range(n_cols2):
            if i == 0:  # 날짜 → 행
                pf = ('<pivotField axis="axisRow" compact="0" outline="0" '
                      'subtotalTop="0" showAll="0" includeNewItemsInFilter="1">'
                      '<items count="1"><item t="default"/></items></pivotField>')
            elif i == 2:  # 월 → 열
                pf = ('<pivotField axis="axisCol" compact="0" outline="0" '
                      'subtotalTop="0" showAll="0" includeNewItemsInFilter="1">'
                      '<items count="1"><item t="default"/></items></pivotField>')
            elif i == val_idx:  # Data → 값
                pf = ('<pivotField dataField="1" compact="0" outline="0" '
                      'subtotalTop="0" showAll="0" includeNewItemsInFilter="1">'
                      '<items count="1"><item t="default"/></items></pivotField>')
            else:
                pf = ('<pivotField compact="0" outline="0" subtotalTop="0" '
                      'showAll="0" includeNewItemsInFilter="1"/>')
            pf_items.append(pf)

        cache_def2 = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
  r:id="rId1" refreshedBy="Python" refreshOnLoad="1" createdVersion="5"
  refreshedVersion="5" minRefreshableVersion="5" recordCount="0">
  <cacheSource type="worksheet">
    <worksheetSource ref="{table2_ref}" sheet="원본 데이터2"/>
  </cacheSource>
  <cacheFields count="{n_cols2}">
{cache_fields_xml}
  </cacheFields>
</pivotCacheDefinition>'''

        cache_rec2 = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                      '<pivotCacheRecords xmlns="http://schemas.openxmlformats.org'
                      '/spreadsheetml/2006/main" count="0"/>')

        cache_rels2 = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                       '<Relationships xmlns="http://schemas.openxmlformats.org'
                       '/package/2006/relationships">\n'
                       '<Relationship Id="rId1" '
                       'Type="http://schemas.openxmlformats.org/officeDocument'
                       '/2006/relationships/pivotCacheRecords" '
                       'Target="pivotCacheRecords2.xml"/>\n'
                       '</Relationships>')

        pivot_tbl2 = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  name="피벗 테이블2" cacheId="{CACHE_ID}"
  applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0"
  applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1"
  dataCaption="값" updatedVersion="5" minRefreshableVersion="5"
  showDrill="1" useAutoFormatting="1" itemPrintTitles="1"
  createdVersion="5" indent="2" compact="0" outline="0"
  outlineData="0" multipleFieldFilters="0">
  <location ref="A3" firstHeaderRow="1" firstDataRow="2" firstDataCol="0"
    rowPageCount="1" colPageCount="1"/>
  <pivotFields count="{n_cols2}">
{"".join(pf_items)}
  </pivotFields>
  <rowFields count="1"><field x="0"/></rowFields>
  <colFields count="1"><field x="2"/></colFields>
  <dataFields count="1">
    <dataField name="합계 : Data" fld="{val_idx}"
      subtotal="sum" showDataAs="normal" baseField="0" baseItem="0"/>
  </dataFields>
</pivotTableDefinition>'''

        pivot_rels2 = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                       '<Relationships xmlns="http://schemas.openxmlformats.org'
                       '/package/2006/relationships">\n'
                       '<Relationship Id="rId1" '
                       'Type="http://schemas.openxmlformats.org/officeDocument'
                       '/2006/relationships/pivotCacheDefinition" '
                       'Target="../pivotCache/pivotCacheDefinition2.xml"/>\n'
                       '</Relationships>')

        # [Content_Types] 수정
        ct = files['[Content_Types].xml'].decode('utf-8')
        new_ct = (
            '<Override PartName="/xl/pivotCache/pivotCacheDefinition2.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.pivotCacheDefinition+xml"/>\n'
            '<Override PartName="/xl/pivotCache/pivotCacheRecords2.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.pivotCacheRecords+xml"/>\n'
            '<Override PartName="/xl/pivotTables/pivotTable2.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.pivotTable+xml"/>'
        )
        ct = ct.replace('</Types>', new_ct + '\n</Types>')
        files['[Content_Types].xml'] = ct.encode('utf-8')

        # workbook.xml에 pivotCaches 추가
        wb_xml = files['xl/workbook.xml'].decode('utf-8')
        new_cache_elem = (f'<pivotCache cacheId="{CACHE_ID}" '
                          f'xmlns:r="http://schemas.openxmlformats.org/officeDocument'
                          f'/2006/relationships" '
                          f'r:id="{CACHE_RID}"/>')
        if '<pivotCaches>' in wb_xml:
            wb_xml = wb_xml.replace('</pivotCaches>',
                                     new_cache_elem + '</pivotCaches>')
        else:
            wb_xml = wb_xml.replace('</workbook>',
                                     f'<pivotCaches>{new_cache_elem}</pivotCaches></workbook>')
        files['xl/workbook.xml'] = wb_xml.encode('utf-8')

        # workbook.xml.rels 수정
        wb_rels = files['xl/_rels/workbook.xml.rels'].decode('utf-8')
        new_rel2 = (f'<Relationship Id="{CACHE_RID}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument'
                    f'/2006/relationships/pivotCacheDefinition" '
                    f'Target="pivotCache/pivotCacheDefinition2.xml"/>')
        wb_rels = wb_rels.replace('</Relationships>', new_rel2 + '</Relationships>')
        files['xl/_rels/workbook.xml.rels'] = wb_rels.encode('utf-8')

        # 피벗작업2 시트 rels
        sheet_fname = pivot2_target.replace('worksheets/', '')
        rels2_key   = f'xl/worksheets/_rels/{sheet_fname}.rels'
        sr2 = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
               '<Relationships xmlns="http://schemas.openxmlformats.org'
               '/package/2006/relationships">\n'
               '<Relationship Id="rIdPivotTable2" '
               'Type="http://schemas.openxmlformats.org/officeDocument'
               '/2006/relationships/pivotTable" '
               'Target="../pivotTables/pivotTable2.xml"/>'
               '\n</Relationships>')
        files[rels2_key] = sr2.encode('utf-8')

        # 새 파일 추가
        files['xl/pivotCache/pivotCacheDefinition2.xml'] = cache_def2.encode('utf-8')
        files['xl/pivotCache/pivotCacheRecords2.xml']    = cache_rec2.encode('utf-8')
        files['xl/pivotCache/_rels/pivotCacheDefinition2.xml.rels'] = cache_rels2.encode('utf-8')
        files['xl/pivotTables/pivotTable2.xml']          = pivot_tbl2.encode('utf-8')
        files['xl/pivotTables/_rels/pivotTable2.xml.rels'] = pivot_rels2.encode('utf-8')

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in files.items():
                zout.writestr(name, data)
        with open(filepath, 'wb') as f:
            f.write(buf.getvalue())

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Box and Whisker (Box Plot) 분석 시트
    # 기온·강수량의 월별·연도별 분포를 시각화
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _sheet_boxplot(self, wb, df: pd.DataFrame, selected_cols: list):
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm
        import matplotlib.patches as mpatches
        import io
        from openpyxl.drawing.image import Image as XLImage

        ws = wb.create_sheet("📦 Box Plot 분석")
        ws.sheet_view.showGridLines = False

        # ── 한글 폰트 설정 ──
        FONT_CANDIDATES = [
            '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
            r'C:\Windows\Fonts\malgun.ttf',
            '/Library/Fonts/NanumGothic.ttf',
        ]
        font_path = None
        for fp in FONT_CANDIDATES:
            import os
            if os.path.exists(fp):
                font_path = fp
                break

        if font_path:
            prop      = fm.FontProperties(fname=font_path)
            font_name = prop.get_name()
            plt.rcParams['font.family'] = font_name
        else:
            prop      = fm.FontProperties()
            font_name = 'sans-serif'
        plt.rcParams['axes.unicode_minus'] = False

        # ── 스타일 설정 ──
        COLORS = {
            'box':      '#2E75B6',
            'median':   '#C55A11',
            'whisker':  '#2E75B6',
            'flier':    '#888888',
            'mean':     '#FF0000',
            'fill':     '#DEEAF1',
        }

        def make_boxplot(data_list, tick_labels, title, ylabel,
                         figsize=(13, 6), color=COLORS['box'],
                         show_mean=True, rot=0):
            """
            data_list : list of 1D arrays (각 박스의 데이터)
            tick_labels: x축 레이블 리스트
            """
            fig, ax = plt.subplots(figsize=figsize)
            fig.patch.set_facecolor('#FAFAFA')
            ax.set_facecolor('#FAFAFA')

            bp = ax.boxplot(
                data_list,
                tick_labels=tick_labels,
                patch_artist=True,
                showmeans=show_mean,
                meanprops=dict(marker='D', markerfacecolor=COLORS['mean'],
                               markeredgecolor=COLORS['mean'], markersize=5),
                flierprops=dict(marker='o', markerfacecolor=COLORS['flier'],
                                markeredgecolor=COLORS['flier'],
                                markersize=3, alpha=0.5),
                medianprops=dict(color=COLORS['median'], linewidth=2.5),
                whiskerprops=dict(color=color, linewidth=1.2),
                capprops=dict(color=color, linewidth=1.5),
            )
            for patch in bp['boxes']:
                patch.set(facecolor=COLORS['fill'], edgecolor=color,
                          linewidth=1.5, alpha=0.85)

            ax.set_title(title, fontproperties=prop, fontsize=13,
                         fontweight='bold', pad=12, color='#1F4E79')
            ax.set_ylabel(ylabel, fontproperties=prop, fontsize=11, color='#333333')
            ax.tick_params(axis='both', labelsize=9)
            if rot:
                ax.set_xticklabels(tick_labels, rotation=rot, ha='right',
                                   fontproperties=prop, fontsize=9)
            else:
                ax.set_xticklabels(tick_labels, fontproperties=prop, fontsize=9)
            ax.yaxis.grid(True, linestyle='--', alpha=0.6, color='#CCCCCC')
            ax.set_axisbelow(True)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            # 범례
            legend_items = [
                mpatches.Patch(facecolor=COLORS['fill'], edgecolor=color, label='IQR (25~75%)'),
                plt.Line2D([0], [0], color=COLORS['median'], linewidth=2.5, label='중앙값'),
            ]
            if show_mean:
                legend_items.append(
                    plt.Line2D([0], [0], marker='D', color='w',
                               markerfacecolor=COLORS['mean'], markersize=7, label='평균')
                )
            ax.legend(handles=legend_items, prop=prop, fontsize=8,
                      loc='upper right', framealpha=0.8)

            plt.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=140, bbox_inches='tight',
                        facecolor='#FAFAFA')
            plt.close(fig)
            buf.seek(0)
            return buf

        def embed_image(ws, buf, anchor, width_px=780, height_px=380):
            img = XLImage(buf)
            img.width  = width_px
            img.height = height_px
            ws.add_image(img, anchor)

        df2 = df.copy()
        df2['_m'] = df2['date'].dt.month
        df2['_y'] = df2['date'].dt.year
        years      = sorted(df2['_y'].unique())
        MONTH_LBLS = ['1월','2월','3월','4월','5월','6월',
                      '7월','8월','9월','10월','11월','12월']

        # ── 시트 제목 ──
        ROW = 1
        _title(ws, ROW, 1, 14,
               'Box and Whisker Plot  —  기온·강수량 월별·연도별 분포 분석',
               h=28)
        ROW += 1
        _note(ws, ROW, 1, 14,
              '※ Box: IQR (25~75%), 중앙선: 중앙값, ◇: 평균, 수염: 1.5×IQR 범위, 점: 이상값')
        ROW += 2

        # 이미지 삽입 행 위치 (엑셀 행 높이 조정)
        for r in range(ROW, ROW + 80):
            ws.row_dimensions[r].height = 15
        for c in range(1, 30):
            ws.column_dimensions[get_column_letter(c)].width = 5.5

        IMG_H = 360   # px
        IMG_W = 780   # px
        ROW_OFFSET = 24   # 이미지 한 장당 행 수

        # ══════════════════════════════════════════════════════
        # 기온 Box Plot (데이터 있을 때만)
        # ══════════════════════════════════════════════════════
        if 'temp_avg' in df.columns and 'temp_avg' in selected_cols:

            # ── ① 평균기온 월별 Box Plot ──
            _title(ws, ROW, 1, 14, '【평균기온】 월별 분포', h=20, bg=C['mid_blue'])
            ROW += 1

            monthly_temp = [
                df2[df2['_m'] == m]['temp_avg'].dropna().values
                for m in range(1, 13)
            ]
            buf = make_boxplot(
                monthly_temp, MONTH_LBLS,
                '월별 평균기온 분포 (℃)\n(각 박스 = 전체 관측기간 해당 월의 일별 데이터)',
                '기온 (℃)',
            )
            embed_image(ws, buf, f'A{ROW}', IMG_W, IMG_H)
            ROW += ROW_OFFSET

            # ── ② 평균기온 연도별 Box Plot ──
            _title(ws, ROW, 1, 14, '【평균기온】 연도별 분포', h=20, bg=C['mid_blue'])
            ROW += 1

            yearly_temp = [
                df2[df2['_y'] == yr]['temp_avg'].dropna().values
                for yr in years
            ]
            yr_lbls = [str(y) for y in years]
            n_yr    = len(years)
            fw      = max(13, n_yr * 0.55)
            buf2 = make_boxplot(
                yearly_temp, yr_lbls,
                '연도별 평균기온 분포 (℃)\n(각 박스 = 해당 연도의 일별 데이터)',
                '기온 (℃)',
                figsize=(fw, 6), rot=45 if n_yr > 15 else 0,
            )
            embed_image(ws, buf2, f'A{ROW}',
                        min(1560, max(780, n_yr * 35)), IMG_H)
            ROW += ROW_OFFSET

        # ── 최고/최저기온 월별 (있을 때) ──
        if ('temp_max' in df.columns and 'temp_max' in selected_cols and
                'temp_min' in df.columns and 'temp_min' in selected_cols):

            _title(ws, ROW, 1, 14, '【최고·최저기온】 월별 분포 비교', h=20, bg=C['mid_blue'])
            ROW += 1

            fig, axes = plt.subplots(1, 2, figsize=(14, 6), facecolor='#FAFAFA')
            fig.suptitle('월별 최고·최저기온 분포 (℃)', fontproperties=prop,
                         fontsize=13, fontweight='bold', color='#1F4E79')

            for ax, col, lbl, clr in [
                (axes[0], 'temp_max', '최고기온 (℃)', '#ED7D31'),
                (axes[1], 'temp_min', '최저기온 (℃)', '#4472C4'),
            ]:
                data_m = [df2[df2['_m'] == m][col].dropna().values for m in range(1, 13)]
                ax.set_facecolor('#FAFAFA')
                bp = ax.boxplot(
                    data_m, tick_labels=MONTH_LBLS,
                    patch_artist=True, showmeans=True,
                    meanprops=dict(marker='D', markerfacecolor='#FF0000',
                                   markeredgecolor='#FF0000', markersize=4),
                    medianprops=dict(color='#C55A11', linewidth=2.5),
                    whiskerprops=dict(color=clr, linewidth=1.2),
                    capprops=dict(color=clr, linewidth=1.5),
                    flierprops=dict(marker='o', markerfacecolor='#888888',
                                    markersize=3, alpha=0.5),
                )
                for patch in bp['boxes']:
                    patch.set(facecolor='#DEEAF1', edgecolor=clr,
                              linewidth=1.5, alpha=0.85)
                ax.set_title(lbl, fontproperties=prop, fontsize=11, color='#1F4E79')
                ax.set_xticklabels(MONTH_LBLS, fontproperties=prop, fontsize=8)
                ax.yaxis.grid(True, linestyle='--', alpha=0.5, color='#CCCCCC')
                ax.set_axisbelow(True)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)

            plt.tight_layout()
            buf3 = io.BytesIO()
            fig.savefig(buf3, format='png', dpi=140, bbox_inches='tight',
                        facecolor='#FAFAFA')
            plt.close(fig)
            buf3.seek(0)
            embed_image(ws, buf3, f'A{ROW}', 1200, IMG_H)
            ROW += ROW_OFFSET

        # ══════════════════════════════════════════════════════
        # 강수량 Box Plot (데이터 있을 때만)
        # ══════════════════════════════════════════════════════
        if 'precipitation' in df.columns and 'precipitation' in selected_cols:

            # ── ③ 강수량 월별 Box Plot ──
            _title(ws, ROW, 1, 14, '【강수량】 월별 분포', h=20, bg=C['mid_blue'])
            ROW += 1

            # 강수량: 0값 제외한 유강우일 데이터 + 전체 데이터 두 버전
            monthly_precip_all = [
                df2[df2['_m'] == m]['precipitation'].dropna().values
                for m in range(1, 13)
            ]
            monthly_precip_wet = [
                df2[(df2['_m'] == m) & (df2['precipitation'] > 0)]['precipitation'].dropna().values
                for m in range(1, 13)
            ]

            fig, axes = plt.subplots(1, 2, figsize=(16, 6), facecolor='#FAFAFA')
            fig.suptitle('월별 강수량 분포 (mm)', fontproperties=prop,
                         fontsize=13, fontweight='bold', color='#1F4E79')

            for ax, data_m, subtitle in [
                (axes[0], monthly_precip_all, '전체(무강우 포함)'),
                (axes[1], monthly_precip_wet,  '유강우일만 (>0mm)'),
            ]:
                ax.set_facecolor('#FAFAFA')
                bp = ax.boxplot(
                    data_m, tick_labels=MONTH_LBLS,
                    patch_artist=True, showmeans=True,
                    meanprops=dict(marker='D', markerfacecolor='#FF0000',
                                   markeredgecolor='#FF0000', markersize=4),
                    medianprops=dict(color='#C55A11', linewidth=2.5),
                    whiskerprops=dict(color='#2E75B6', linewidth=1.2),
                    capprops=dict(color='#2E75B6', linewidth=1.5),
                    flierprops=dict(marker='o', markerfacecolor='#888888',
                                    markersize=2, alpha=0.4),
                )
                for patch in bp['boxes']:
                    patch.set(facecolor='#DEEAF1', edgecolor='#2E75B6',
                              linewidth=1.5, alpha=0.85)
                ax.set_title(subtitle, fontproperties=prop, fontsize=11, color='#1F4E79')
                ax.set_ylabel('강수량 (mm)', fontproperties=prop, fontsize=10)
                ax.set_xticklabels(MONTH_LBLS, fontproperties=prop, fontsize=8)
                ax.yaxis.grid(True, linestyle='--', alpha=0.5, color='#CCCCCC')
                ax.set_axisbelow(True)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)

            plt.tight_layout()
            buf4 = io.BytesIO()
            fig.savefig(buf4, format='png', dpi=140, bbox_inches='tight',
                        facecolor='#FAFAFA')
            plt.close(fig)
            buf4.seek(0)
            embed_image(ws, buf4, f'A{ROW}', 1300, IMG_H)
            ROW += ROW_OFFSET

            # ── ④ 강수량 연도별 Box Plot ──
            _title(ws, ROW, 1, 14, '【강수량】 연도별 분포 (유강우일, >0mm)', h=20, bg=C['mid_blue'])
            ROW += 1

            yearly_precip = [
                df2[(df2['_y'] == yr) & (df2['precipitation'] > 0)]['precipitation'].dropna().values
                for yr in years
            ]
            fw = max(13, n_yr * 0.55)
            buf5 = make_boxplot(
                yearly_precip, yr_lbls,
                '연도별 강수량 분포 (mm, 유강우일 기준)\n(각 박스 = 해당 연도의 유강우일 일강수량)',
                '강수량 (mm)',
                figsize=(fw, 6),
                color='#2E75B6',
                rot=45 if n_yr > 15 else 0,
            )
            embed_image(ws, buf5, f'A{ROW}',
                        min(1560, max(780, n_yr * 35)), IMG_H)
            ROW += ROW_OFFSET

        # ── 통계 요약 표 ──
        ROW += 1
        _title(ws, ROW, 1, 14, '통계 요약 — 월별 사분위수 (IQR)', h=20, bg=C['mid_blue'])
        ROW += 1

        stat_cols_to_show = []
        if 'temp_avg' in df.columns and 'temp_avg' in selected_cols:
            stat_cols_to_show.append(('temp_avg', '평균기온(℃)'))
        if 'precipitation' in df.columns and 'precipitation' in selected_cols:
            stat_cols_to_show.append(('precipitation', '강수량(mm)'))

        for ck, lbl in stat_cols_to_show:
            ws.merge_cells(start_row=ROW, start_column=1,
                           end_row=ROW, end_column=8)
            c = ws.cell(row=ROW, column=1,
                        value=f'[ {lbl} ] 월별 통계 (Q1·중앙값·Q3·평균·이상값 기준)')
            c.font      = Font(name=FONT, bold=True, size=10, color=C['dark_blue'])
            c.alignment = Alignment(horizontal='left', indent=1)
            ws.row_dimensions[ROW].height = 16
            ROW += 1

            # 헤더
            for j, h in enumerate(['월','최소','Q1(25%)','중앙값','평균','Q3(75%)','최대','IQR'], 1):
                _hc(ws, ROW, j, h, bg=C['dark_blue'], sz=9)
                ws.column_dimensions[get_column_letter(j)].width = 11
            ws.row_dimensions[ROW].height = 16
            ROW += 1

            for m in range(1, 13):
                ser = df2[df2['_m'] == m][ck].dropna()
                if ser.empty:
                    continue
                q1  = round(float(np.percentile(ser, 25)), 1)
                med = round(float(np.percentile(ser, 50)), 1)
                q3  = round(float(np.percentile(ser, 75)), 1)
                avg = round(float(ser.mean()), 1)
                mn  = round(float(ser.min()), 1)
                mx  = round(float(ser.max()), 1)
                iqr = round(q3 - q1, 1)
                bg  = C['light_blue'] if m % 2 == 0 else C['white']
                for j, v in enumerate([f'{m}월', mn, q1, med, avg, q3, mx, iqr], 1):
                    _dc(ws, ROW, j, v, bg=bg, sz=9,
                        nf='#,##0.0' if isinstance(v, float) else None)
                ws.row_dimensions[ROW].height = 15
                ROW += 1
            ROW += 1
